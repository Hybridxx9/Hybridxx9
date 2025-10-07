import logging
import asyncio
import time
import os
import csv
import traceback
import openpyxl
from urllib.parse import urlencode
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv
from io import BytesIO, StringIO
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
import aiohttp
import json

# Загрузка переменных окружения
load_dotenv()

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==== КОНФИГУРАЦИЯ ИЗ ПЕРЕМЕННЫХ ОКРУЖЕНИЯ ====
API_TOKEN = os.getenv("BOT_TOKEN", "8342338980:AAFqW9vU1pT0Rwa8y7Z7vL0yiJ6nD99dHNQ")
ADMIN_ID = int(os.getenv("ADMIN_ID", "436919904"))

# Получаем список разрешенных пользователей из переменной окружения
allowed_users_str = os.getenv("ALLOWED_USERS", "")
ALLOWED_USERS = [int(user_id.strip()) for user_id in allowed_users_str.split(",") if user_id.strip()]
# Добавляем стандартных пользователей если в переменной пусто
if not ALLOWED_USERS:
    ALLOWED_USERS = [436919904, 8153905248, 438098732]
# Добавляем ADMIN_ID в список если его там нет
if ADMIN_ID not in ALLOWED_USERS:
    ALLOWED_USERS.append(ADMIN_ID)

ETHERSCAN_API_KEY = os.getenv("ETHERSCAN_API_KEY", "46N2KJPSTIUJ43TKT7IR23YKGZ35WRAEST")

# Проверяем обязательные переменные
if not API_TOKEN:
    raise ValueError("BOT_TOKEN не установлен")
if not ADMIN_ID:
    raise ValueError("ADMIN_ID не установлен")
if not ETHERSCAN_API_KEY:
    raise ValueError("ETHERSCAN_API_KEY не установлен")

# Контракты токенов
USDC_CONTRACT_BASE = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
USDC_CONTRACT_ARBITRUM = "0xaf88d065e77c8cC2239327C5EDb3A432268e5831" 
USDC_CONTRACT_OPTIMISM = "0x0b2C639c533813f4Aa9D7837CAf62653d097Ff85"

ETH_CONTRACT = "0x0000000000000000000000000000000000000000"

# API конфигурация - ЕДИНЫЙ ENDPOINT
ETHERSCAN_V2_API_URL = "https://api.etherscan.io/v2/api"

# Basescan API
BASESCAN_API_URL = "https://api.basescan.org/api"

BASE_RPC_URL = "https://mainnet.base.org"

# Chain IDs
BASE_CHAIN_ID = 8453
ARBITRUM_CHAIN_ID = 42161  
OPTIMISM_CHAIN_ID = 10

# RPC URLs
BASE_RPC_URL = "https://mainnet.base.org"
ARBITRUM_RPC_URL = "https://arb1.arbitrum.io/rpc"

# Приоритет сетей
NETWORK_PRIORITY = [
    {"name": "Base", "chain_id": BASE_CHAIN_ID, "usdc_contract": USDC_CONTRACT_BASE},
    {"name": "Arbitrum", "chain_id": ARBITRUM_CHAIN_ID, "usdc_contract": USDC_CONTRACT_ARBITRUM},
    {"name": "Optimism", "chain_id": OPTIMISM_CHAIN_ID, "usdc_contract": USDC_CONTRACT_OPTIMISM}
]

# ========== ИНИЦИАЛИЗАЦИЯ БОТА ==========
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Состояния FSM
class AnalyzeState(StatesGroup):
    waiting_for_contract = State()
    waiting_for_mode = State()
    waiting_for_token = State()
    waiting_for_network = State()

class AllowanceState(StatesGroup):
    waiting_for_source = State()
    waiting_for_wallets = State()
    waiting_for_token_allowance = State()
    waiting_for_spender = State()

# Глобальные переменные
contract_analysis_results = {}
user_sessions = {}
last_user_sessions = {}  # Для отслеживания последней сессии пользователя
# Глобальная переменная для хранения текущих данных allowance
current_allowance_data = {}

# ========== КЛАВИАТУРЫ ==========

def get_admin_menu_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🔍 Анализ контракта")],
            [KeyboardButton(text="💰 Проверка allowance")],
            [KeyboardButton(text="🛠️ Диагностика")]
        ],
        resize_keyboard=True,
        persistent=True
    )

def get_network_selection_keyboard_allowance():
    """Клавиатура для выбора сети при проверке allowance"""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔵 Base", callback_data="allowance_network_base")],
        [InlineKeyboardButton(text="🔷 Arbitrum", callback_data="allowance_network_arbitrum")],
        [InlineKeyboardButton(text="🟠 Optimism", callback_data="allowance_network_optimism")],
        [InlineKeyboardButton(text="🌐 Все сети", callback_data="allowance_network_all")]
    ])

def get_analysis_mode_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔍 Одна сеть", callback_data="mode_single")],
        [InlineKeyboardButton(text="🌐 Все сети", callback_data="mode_all")]
    ])

def get_token_selection_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💙 USDC", callback_data="token_usdc")],
        [InlineKeyboardButton(text="🔷 ETH (❌НЕ ДЛЯ allowance ❌ )", callback_data="token_eth")]
    ])

def get_network_selection_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔵 Base", callback_data="network_base")],
        [InlineKeyboardButton(text="🔷 Arbitrum", callback_data="network_arbitrum")],
        [InlineKeyboardButton(text="🟠 Optimism", callback_data="network_optimism")]
    ])

def get_allowance_source_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Использовать результат анализа", callback_data="source_analysis")],
        [InlineKeyboardButton(text="📁 Загрузить новый список", callback_data="source_upload")]
    ])

# ========== УТИЛИТЫ ДЛЯ API ==========

async def make_etherscan_request(chain_id, params):
    """Универсальный запрос к Etherscan V2 API - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
    try:
        # Добавляем chainid в параметры
        params['chainid'] = chain_id
        params['apikey'] = ETHERSCAN_API_KEY
        
        # Базовый URL без chainid
        api_url = "https://api.etherscan.io/v2/api"
        
        logger.info(f"🔧 V2 API Request: {api_url}, params={params}")
        
        async with aiohttp.ClientSession() as session:
            async with session.get(api_url, params=params, timeout=30) as response:
                if response.status == 200:
                    data = await response.json()
                    
                    logger.info(f"🔧 V2 API Response: status={data.get('status')}, message={data.get('message')}, result={data.get('result')}")
                    
                    if data.get('status') == '1' or data.get('message') == 'OK':
                        return data
                    else:
                        logger.error(f"❌ API Error: {data.get('message')}")
                        return None
                else:
                    logger.error(f"❌ HTTP Error: {response.status}")
                    return None
    except Exception as e:
        logger.error(f"❌ API Request Error: {e}")
        return None

async def make_basescan_request(params):
    """Прямой запрос через Etherscan V2 API для Base сети"""
    try:
        params['chainid'] = BASE_CHAIN_ID
        params['apikey'] = ETHERSCAN_API_KEY
        
        api_url = "https://api.etherscan.io/v2/api"
        
        logger.info(f"🔧 [BASESCAN_V2] Request: {api_url}, params={params}")
        
        async with aiohttp.ClientSession() as session:
            async with session.get(api_url, params=params, timeout=30) as response:
                if response.status == 200:
                    data = await response.json()
                    logger.info(f"🔧 [BASESCAN_V2] Response: status={data.get('status')}, message={data.get('message')}")
                    
                    if data.get('status') == '1' or data.get('message') == 'OK':
                        return data
                    else:
                        logger.error(f"❌ [BASESCAN_V2] API Error: {data.get('message')}")
                        return None
                else:
                    logger.error(f"❌ [BASESCAN_V2] HTTP Error: {response.status}")
                    return None
    except Exception as e:
        logger.error(f"❌ [BASESCAN_V2] Request Error: {e}")
        return None

async def get_contract_transactions(contract_address, chain_id, token_contract=None):
    """Версия для точного соответствия с сайтом Arbitrum Scan"""
    if token_contract:
        params = {
            'module': 'account',
            'action': 'tokentx',
            'contractaddress': token_contract,
            'address': contract_address,
            'sort': 'desc'
        }
    else:
        params = {
            'module': 'account',
            'action': 'txlist',
            'address': contract_address,
            'sort': 'desc'
        }
    
    all_transactions = []
    page = 1
    unique_address_pairs = set()
    
    # ⚠️ ПОЛУЧАЕМ ВСЕ СТРАНИЦЫ ДО КОНЦА
    while True:
        params['page'] = page
        params['offset'] = 1000
        
        logger.info(f"📄 Запрос страницы {page}")
        
        data = await make_etherscan_request(chain_id, params)
        
        if not data or not isinstance(data.get('result'), list):
            break
            
        transactions = data['result']
        
        if not transactions:
            break
        
        # Фильтруем дубликаты
        for tx in transactions:
            from_addr = tx.get('from', '').lower()
            to_addr = tx.get('to', '').lower()
            address_pair = (from_addr, to_addr)
            
            if from_addr and to_addr and address_pair not in unique_address_pairs:
                unique_address_pairs.add(address_pair)
                all_transactions.append(tx)
        
        current_total = len(all_transactions)
        logger.info(f"📥 Страница {page}: {len(transactions)} записей, {current_total} уникальных пар")
        
        # ⚠️ КРИТИЧЕСКОЕ ИЗМЕНЕНИЕ: ПРЕКРАЩАЕМ ТОЛЬКО КОГДА ПОЛУЧАЕМ ПУСТУЮ СТРАНИЦУ
        if len(transactions) < 1000:
            logger.info(f"✅ Получены все транзакции. Последняя страница: {page}")
            break
            
        page += 1
        
        # ⚠️ ЗАЩИТА ОТ БЕСКОНЕЧНОГО ЦИКЛА
        if page > 10:
            logger.warning("⚠️ Достигнут лимит в 10 страниц")
            break
            
        await asyncio.sleep(0.1)
    
    total_count = len(all_transactions)
    logger.info(f"📊 ФИНАЛЬНЫЙ ИТОГ: {total_count} уникальных транзакций")
    
    return all_transactions, total_count

async def get_contract_transactions_complete(contract_address, chain_id, token_contract=None):
    """Полная версия - получает ВСЕ транзакции для нахождения ВСЕХ кошельков"""
    if token_contract:
        params = {
            'module': 'account',
            'action': 'tokentx',
            'contractaddress': token_contract,
            'address': contract_address,
            'sort': 'desc'
        }
    else:
        params = {
            'module': 'account',
            'action': 'txlist',
            'address': contract_address,
            'sort': 'desc'
        }
    
    all_transactions = []
    page = 1
    
    # ⚠️ БЕЗ ОГРАНИЧЕНИЙ - ПОЛУЧАЕМ ВСЕ СТРАНИЦЫ
    while True:
        params['page'] = page
        params['offset'] = 1000
        
        logger.info(f"📄 Запрос страницы {page} для полного анализа")
        
        data = await make_etherscan_request(chain_id, params)
        
        if not data:
            logger.warning(f"⚠️ Нет данных на странице {page}")
            break
            
        result = data.get('result')
        
        # Обрабатываем ошибки API
        if isinstance(result, str):
            if 'no transactions' in result.lower():
                logger.info(f"📭 Больше транзакций нет")
                break
            elif any(error in result.lower() for error in ['error', 'max', 'rate limit']):
                logger.error(f"❌ API ошибка: {result}")
                break
        
        if not isinstance(result, list):
            logger.error(f"❌ Ожидался список, получен: {type(result)}")
            break
            
        transactions = result
        
        if not transactions:
            logger.info(f"📭 Больше транзакций нет на странице {page}")
            break
        
        # ⚠️ ФИЛЬТРУЕМ ТОЛЬКО ПО ХЕШАМ (чтобы избежать дубликатов одной транзакции)
        unique_hashes = set()
        new_transactions = []
        
        for tx in transactions:
            tx_hash = tx.get('hash')
            if tx_hash and tx_hash not in unique_hashes:
                unique_hashes.add(tx_hash)
                new_transactions.append(tx)
        
        all_transactions.extend(new_transactions)
        current_total = len(all_transactions)
        
        logger.info(f"📥 Страница {page}: {len(transactions)} записей, {len(new_transactions)} уникальных хешей, всего: {current_total}")
        
        # Логируем прогресс каждые 5 страниц
        if page % 5 == 0:
            logger.info(f"🔄 Обработано {page} страниц, найдено {current_total} транзакций")
        
        # ⚠️ ПРЕКРАЩАЕМ ТОЛЬКО КОГДА API ВОЗВРАЩАЕТ МЕНЬШЕ 1000 ТРАНЗАКЦИЙ
        if len(transactions) < 1000:
            logger.info(f"✅ Получены ВСЕ транзакции контракта. Последняя страница: {page}")
            break
            
        page += 1
        
        # Защита от бесконечного цикла (но с большим запасом)
        if page > 50:
            logger.warning(f"⚠️ Достигнут лимит в 50 страниц. Получено {current_total} транзакций")
            break
            
        await asyncio.sleep(0.3)
    
    total_count = len(all_transactions)
    logger.info(f"📊 ПОЛНЫЙ ИТОГ: {total_count} транзакций, {page} страниц")
    
    return all_transactions, total_count

async def get_contract_transactions_accurate(contract_address, chain_id, token_contract=None):
    """Точная версия, которая фильтрует только основные транзакции"""
    if token_contract:
        params = {
            'module': 'account',
            'action': 'tokentx',
            'contractaddress': token_contract,
            'address': contract_address,
            'sort': 'desc'
        }
    else:
        params = {
            'module': 'account',
            'action': 'txlist',
            'address': contract_address,
            'sort': 'desc'
        }
    
    all_transactions = []
    page = 1
    
    # ⚠️ ОТСЛЕЖИВАЕМ УНИКАЛЬНЫЕ ХЕШИ ТРАНЗАКЦИЙ (а не пары адресов)
    unique_hashes = set()
    
    while True:
        params['page'] = page
        params['offset'] = 1000
        
        logger.info(f"📄 Запрос страницы {page} для точного анализа")
        
        data = await make_etherscan_request(chain_id, params)
        
        if not data or not isinstance(data.get('result'), list):
            break
            
        transactions = data['result']
        
        if not transactions:
            break
        
        # ⚠️ КРИТИЧЕСКОЕ ИЗМЕНЕНИЕ: ФИЛЬТРУЕМ ПО УНИКАЛЬНЫМ ХЕШАМ
        new_transactions = []
        for tx in transactions:
            tx_hash = tx.get('hash')
            if tx_hash and tx_hash not in unique_hashes:
                unique_hashes.add(tx_hash)
                new_transactions.append(tx)
        
        all_transactions.extend(new_transactions)
        current_total = len(all_transactions)
        
        logger.info(f"📥 Страница {page}: {len(transactions)} записей, {len(new_transactions)} уникальных хешей, всего: {current_total}")
        
        if page == 1 and new_transactions:
            for i, tx in enumerate(new_transactions[:3]):
                logger.info(f"🔍 [ACCURATE_SAMPLE_{i}] From: {tx.get('from')}, To: {tx.get('to')}, Hash: {tx.get('hash')}")
        
        # Прекращаем когда получаем меньше транзакций
        if len(transactions) < 1000:
            logger.info(f"✅ Получены все основные транзакции")
            break
            
        page += 1
        
        if page > 10:
            logger.warning("⚠️ Достигнут лимит в 10 страниц")
            break
            
        await asyncio.sleep(0.3)
    
    total_count = len(all_transactions)
    logger.info(f"📊 ТОЧНЫЙ ИТОГ: {total_count} основных транзакций (по хешам)")
    
    return all_transactions, total_count

# ========== КОНСТАНТЫ ДЛЯ КЭШИРОВАНИЯ ==========
PROGRESS_CACHE_FILE = "allowance_progress.json"
CACHE_SAVE_INTERVAL = 25  # Сохранять каждые 25 кошельков
# ========== ФУНКЦИИ КЭШИРОВАНИЯ ПРОГРЕССА ==========

def load_progress(user_id):
    """Загружает прогресс из файла"""
    try:
        with open(PROGRESS_CACHE_FILE, "r") as f:
            all_progress = json.load(f)
            return all_progress.get(str(user_id))
    except (FileNotFoundError, json.JSONDecodeError):
        return None

def save_progress(user_id, progress_data):
    """Сохраняет прогресс в файл"""
    try:
        # Загружаем существующие данные
        try:
            with open(PROGRESS_CACHE_FILE, "r") as f:
                all_progress = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            all_progress = {}
        
        # Обновляем прогресс для пользователя
        all_progress[str(user_id)] = progress_data
        
        # Сохраняем обратно
        with open(PROGRESS_CACHE_FILE, "w") as f:
            json.dump(all_progress, f, indent=2)
            
        logger.info(f"💾 Прогресс сохранен: {progress_data['current_index']}/{progress_data['total_wallets']}")
    except Exception as e:
        logger.error(f"❌ Ошибка сохранения прогресса: {e}")

def delete_progress(user_id):
    """Удаляет прогресс пользователя после завершения"""
    try:
        with open(PROGRESS_CACHE_FILE, "r") as f:
            all_progress = json.load(f)
        
        if str(user_id) in all_progress:
            del all_progress[str(user_id)]
            
        with open(PROGRESS_CACHE_FILE, "w") as f:
            json.dump(all_progress, f, indent=2)
            
        logger.info(f"🧹 Прогресс пользователя {user_id} удален")
    except Exception as e:
        logger.error(f"❌ Ошибка удаления прогресса: {e}")

# ========== БАЗОВЫЕ ФУНКЦИИ ДЛЯ ПРОВЕРКИ БАЛАНСА И ALLOWANCE ==========

async def get_token_balance(wallet_address, token_contract, chain_id, is_eth=False):
    """Получение баланса токена или ETH"""
    try:
        logger.info(f"🔧 [BALANCE_DETAILED] Запрос баланса для {wallet_address}")
        
        if is_eth:
            params = {
                'module': 'account',
                'action': 'balance',
                'address': wallet_address,
                'tag': 'latest'
            }
        else:
            params = {
                'module': 'account',
                'action': 'tokenbalance',
                'contractaddress': token_contract,
                'address': wallet_address,
                'tag': 'latest'
            }
        
        data = await make_etherscan_request(chain_id, params)
        
        logger.info(f"🔧 [BALANCE_DETAILED] Полный ответ API: {data}")
        
        if data and (data.get('status') == '1' or data.get('message') == 'OK'):
            balance_raw = data.get('result', '0')
            logger.info(f"🔧 [BALANCE_DETAILED] Raw баланс: '{balance_raw}' (тип: {type(balance_raw)})")
            
            # 🔧 КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: Проверяем что balance_raw не пустой и не '0'
            if not balance_raw or balance_raw == '0':
                logger.warning(f"🔧 [BALANCE_DETAILED] Raw баланс пустой или '0'")
                return 0
            
            try:
                # Конвертируем в int
                balance_int = int(balance_raw)
                logger.info(f"🔧 [BALANCE_DETAILED] Баланс как int: {balance_int}")
                
                if is_eth:
                    balance = balance_int / (10 ** 18)  # ETH decimals
                else:
                    balance = balance_int / (10 ** 6)   # USDC decimals
                
                logger.info(f"🔧 [BALANCE_DETAILED] Рассчитанный баланс: {balance}")
                return balance
                
            except ValueError as e:
                logger.error(f"❌ Ошибка конвертации баланса '{balance_raw}': {e}")
                return 0
        else:
            error_msg = data.get('message', 'No response') if data else 'No data'
            logger.warning(f"🔧 [BALANCE_DETAILED] Ошибка API: {error_msg}")
            return 0
            
    except Exception as e:
        logger.error(f"❌ Balance check error for {wallet_address[:8]}: {e}")
        return 0

# ========== БАЗОВЫЕ ФУНКЦИИ ДЛЯ ПРОВЕРКИ БАЛАНСА И ALLOWANCE ==========

async def get_token_balance(wallet_address, token_contract, chain_id, is_eth=False):
    """Получение баланса токена или ETH"""
    try:
        if is_eth:
            # Для ETH
            params = {
                'module': 'account',
                'action': 'balance',
                'address': wallet_address,
                'tag': 'latest'
            }
        else:
            # Для ERC20 токенов (USDC)
            params = {
                'module': 'account',
                'action': 'tokenbalance',
                'contractaddress': token_contract,
                'address': wallet_address,
                'tag': 'latest'
            }
        
        data = await make_etherscan_request(chain_id, params)
        
        if data and (data.get('status') == '1' or data.get('message') == 'OK'):
            balance_raw = data.get('result', '0')
            if is_eth:
                balance = int(balance_raw) / (10 ** 18)  # ETH decimals
            else:
                balance = int(balance_raw) / (10 ** 6)   # USDC decimals
            return balance
        else:
            return 0
            
    except Exception as e:
        logger.error(f"❌ Balance check error for {wallet_address[:8]}: {e}")
        return 0

# ========== ФУНКЦИИ С RETRY ЛОГИКОЙ ==========

async def get_token_balance_with_retry(wallet, token_contract, chain_id, is_eth=False, max_retries=3):
    """Проверка баланса с retry логикой и adaptive pacing"""
    base_delay = 1.0  # Базовая задержка
    
    for attempt in range(max_retries):
        try:
            # Теперь эта функция существует!
            balance = await get_token_balance(wallet, token_contract, chain_id, is_eth)
            return balance
            
        except Exception as e:
            error_str = str(e)
            
            # Adaptive pacing: увеличиваем задержку при rate limit
            if "429" in error_str or "rate limit" in error_str or "Too Many Requests" in error_str:
                wait_time = base_delay * (2 ** attempt)  # Экспоненциальная задержка: 1s, 2s, 4s
                logger.warning(f"⚠️ Rate limit для {wallet[:8]}... (попытка {attempt+1}), ждем {wait_time}с")
                await asyncio.sleep(wait_time)
                
                # Увеличиваем базовую задержку на 25% для следующих запросов
                base_delay *= 1.25
                continue
            else:
                logger.error(f"❌ Ошибка баланса для {wallet[:8]}...: {error_str}")
                return 0
    
    logger.error(f"❌ Все попытки не удались для {wallet[:8]}...")
    return 0        


async def get_token_balance_with_retry(wallet, token_contract, chain_id, is_eth=False, max_retries=3):
    """Проверка баланса с retry логикой и adaptive pacing"""
    # КОРРЕКТИРОВКА: Используем правильный контракт USDC для каждой сети
    if not is_eth and token_contract == USDC_CONTRACT_BASE:
        if chain_id == ARBITRUM_CHAIN_ID:
            token_contract = USDC_CONTRACT_ARBITRUM
        elif chain_id == OPTIMISM_CHAIN_ID:
            token_contract = USDC_CONTRACT_OPTIMISM
    
    base_delay = 1.0  # Базовая задержка
    
    for attempt in range(max_retries):
        try:
            balance = await get_token_balance(wallet, token_contract, chain_id, is_eth)
            return balance
            
        except Exception as e:
            error_str = str(e)
            
            # Adaptive pacing: увеличиваем задержку при rate limit
            if "429" in error_str or "rate limit" in error_str or "Too Many Requests" in error_str:
                wait_time = base_delay * (2 ** attempt)  # Экспоненциальная задержка: 1s, 2s, 4s
                logger.warning(f"⚠️ Rate limit для {wallet[:8]}... (попытка {attempt+1}), ждем {wait_time}с")
                await asyncio.sleep(wait_time)
                
                # Увеличиваем базовую задержку на 25% для следующих запросов
                base_delay *= 1.25
                continue
            else:
                logger.error(f"❌ Ошибка баланса для {wallet[:8]}...: {error_str}")
                return 0
    
    logger.error(f"❌ Все попытки не удались для {wallet[:8]}...")
    return 0

async def get_allowance_basescan(wallet_address, token_contract, spender_address):
    """Прямой запрос allowance через V2 API для Base сети"""
    try:
        params = {
            'module': 'account',
            'action': 'tokenallowance',
            'contractaddress': token_contract,
            'address': wallet_address,
            'spender': spender_address
        }
        
        data = await make_basescan_request(params)
        
        if data and (data.get('status') == '1' or data.get('message') == 'OK'):
            allowance_raw = data.get('result', '0')
            allowance = int(allowance_raw)
            allowance_usd = allowance / (10 ** 6)  # USDC decimals
            
            short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
            short_spender = spender_address[:8] + "..." + spender_address[-6:]
            logger.info(f"✅ [BASESCAN_ALLOWANCE_V2] {short_wallet} -> {short_spender}: {allowance_usd:,.6f} USDC")
            
            return allowance_usd
            
    except Exception as e:
        logger.error(f"❌ [BASESCAN_ALLOWANCE_V2_ERROR] {e}")
    
    return 0

async def get_token_allowance(wallet_address, token_contract, spender_address, chain_id):
    """Обновленная функция - использует правильные контракты для каждой сети"""
    # КОРРЕКТИРОВКА: Используем правильный контракт для каждой сети
    if chain_id == ARBITRUM_CHAIN_ID and token_contract == USDC_CONTRACT_BASE:
        token_contract = USDC_CONTRACT_ARBITRUM
    elif chain_id == OPTIMISM_CHAIN_ID and token_contract == USDC_CONTRACT_BASE:
        token_contract = USDC_CONTRACT_OPTIMISM
    
    # Пробуем стандартный V2 метод
    params = {
        'module': 'account',
        'action': 'tokenallowance',
        'contractaddress': token_contract,  # Теперь здесь правильный контракт
        'address': wallet_address,
        'spender': spender_address
    }
    
    data = await make_etherscan_request(chain_id, params)
    
    if data and (data.get('status') == '1' or data.get('message') == 'OK'):
        allowance_raw = data.get('result', '0')
        allowance = int(allowance_raw)
        allowance_usd = allowance / (10 ** 6)
        
        short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
        short_spender = spender_address[:8] + "..." + spender_address[-6:]
        logger.info(f"✅ [STANDARD_ALLOWANCE] {short_wallet} -> {short_spender}: {allowance_usd:,.6f} USDC")
        
        return allowance_usd
    else:
        error_msg = data.get('message', 'No response') if data else 'No data'
        logger.warning(f"⚠️ [STANDARD_ALLOWANCE_FAILED] {wallet_address}: {error_msg}")
        
        return 0
    
async def get_allowance_via_rpc(wallet_address, token_contract, spender_address):
    """Прямой вызов контракта через Base RPC - С ЗАДЕРЖКАМИ"""
    try:
        # Данные для вызова функции allowance(owner, spender)
        data_payload = "0xdd62ed3e" + wallet_address[2:].zfill(64) + spender_address[2:].zfill(64)
        
        # Формируем JSON-RPC запрос
        rpc_payload = {
            "jsonrpc": "2.0",
            "method": "eth_call",
            "params": [{
                "to": token_contract,
                "data": data_payload
            }, "latest"],
            "id": 1
        }
        
        logger.info(f"🔧 [RPC_CALL] Calling contract via Base RPC...")
        
        # ЗАДЕРЖКА ДЛЯ ИЗБЕЖАНИЯ RATE LIMIT
        await asyncio.sleep(0.3)
        
        async with aiohttp.ClientSession() as session:
            async with session.post(BASE_RPC_URL, json=rpc_payload, timeout=30) as response:
                if response.status == 200:
                    data = await response.json()
                    
                    if data.get('result'):
                        result_hex = data['result']
                        if result_hex != '0x':
                            allowance_raw = int(result_hex, 16)
                            
                            # MAX_UINT256 значение
                            MAX_UINT256 = 115792089237316195423570985008687907853269984665640564039457584007913129639935
                            
                            if allowance_raw == MAX_UINT256:
                                # Для MAX_UINT256 возвращаем баланс как бесконечный allowance
                                params = {
                                    'module': 'account',
                                    'action': 'tokenbalance',
                                    'contractaddress': token_contract,
                                    'address': wallet_address,
                                    'tag': 'latest'
                                }
                                balance_data = await make_etherscan_request(BASE_CHAIN_ID, params)
                                if balance_data and balance_data.get('status') == '1':
                                    balance_raw = balance_data.get('result', '0')
                                    balance = int(balance_raw) / (10 ** 6)
                                    
                                    short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
                                    short_spender = spender_address[:8] + "..." + spender_address[-6:]
                                    logger.info(f"🎯 [RPC_MAX_ALLOWANCE] {short_wallet} -> {short_spender}: MAX_UINT256 = {balance:.6f} USDC")
                                    
                                    return balance
                            else:
                                allowance_usd = allowance_raw / (10 ** 6)
                                
                                short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
                                short_spender = spender_address[:8] + "..." + spender_address[-6:]
                                logger.info(f"🎯 [RPC_ALLOWANCE] {short_wallet} -> {short_spender}: {allowance_usd:,.6f} USDC")
                                
                                return allowance_usd
                    
                    logger.warning(f"⚠️ [RPC_CALL] No valid result for {wallet_address}")
                    return 0
                elif response.status == 429:
                    logger.warning(f"⚠️ [RPC_CALL] Rate limited, waiting...")
                    await asyncio.sleep(1)  # Увеличиваем задержку при rate limit
                    return 0
                else:
                    logger.error(f"❌ [RPC_CALL] HTTP Error: {response.status}")
                    return 0
                    
    except Exception as e:
        logger.error(f"❌ [RPC_CALL_ERROR] {e}")
        return 0

async def get_allowance_direct(wallet_address, token_contract, spender_address, chain_id):
    """Прямой вызов контракта через Etherscan API - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
    try:
        # ABI для вызова функции allowance
        data_payload = "0xdd62ed3e"  # allowance(address,address) function selector
        
        # Добавляем параметры (pad to 32 bytes each)
        data_payload += wallet_address[2:].zfill(64)  # owner
        data_payload += spender_address[2:].zfill(64) # spender
        
        params = {
            'module': 'proxy',
            'action': 'eth_call',
            'to': token_contract,
            'data': data_payload,
            'tag': 'latest'
        }
        
        data = await make_etherscan_request(chain_id, params)
        
        if data and data.get('result') and data.get('result') != '0x':
            # Результат в hex, конвертируем в decimal
            allowance_hex = data['result']
            allowance_raw = int(allowance_hex, 16)
            
            # MAX_UINT256 значение
            MAX_UINT256 = 115792089237316195423570985008687907853269984665640564039457584007913129639935
            
            if allowance_raw == MAX_UINT256:
                # Для MAX_UINT256 возвращаем баланс кошелька как доступную сумму
                is_eth = (token_contract == ETH_CONTRACT)
                balance = await get_token_balance(wallet_address, token_contract, chain_id, is_eth)
                allowance_usd = balance  # Весь баланс доступен
                
                short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
                short_spender = spender_address[:8] + "..." + spender_address[-6:]
                logger.info(f"🎯 [MAX_ALLOWANCE] {short_wallet} -> {short_spender}: MAX_UINT256 = весь баланс {balance:.2f} USDC")
                
                return allowance_usd
            else:
                allowance_usd = allowance_raw / (10 ** 6)  # USDC decimals
                
                short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
                short_spender = spender_address[:8] + "..." + spender_address[-6:]
                logger.info(f"🎯 [DIRECT_ALLOWANCE] {short_wallet} -> {short_spender}: {allowance_usd:,.6f} USDC")
                
                return allowance_usd
        else:
            logger.warning(f"⚠️ [DIRECT_ALLOWANCE] No result for {wallet_address}")
            
    except Exception as e:
        logger.error(f"❌ [DIRECT_ALLOWANCE_ERROR] {e}")
    
    return 0

# ========== МОДУЛЬ 1: АНАЛИЗ КОНТРАКТОВ ==========

async def analyze_contract_all_networks(contract_address, token_type, progress_callback=None):
    """Анализ контракта во всех сетях - С ДЕТАЛЬНЫМ ЛОГИРОВАНИЕМ"""
    all_wallets = set()
    total_processed = 0
    
    for i, network in enumerate(NETWORK_PRIORITY):
        if progress_callback:
            progress = int((i / len(NETWORK_PRIORITY)) * 100)
            await progress_callback(f"🔍 Анализ в {network['name']}...", progress)
        
        logger.info(f"🔍 Анализ контракта в {network['name']}")
        
        # Определяем контракт токена
        token_contract = network['usdc_contract'] if token_type == 'usdc' else ETH_CONTRACT
        
        transactions, tx_count = await get_contract_transactions(
            contract_address, 
            network['chain_id'], 
            token_contract if token_type == 'usdc' else None
        )
        
        logger.info(f"🔍 Обработка {len(transactions)} транзакций в {network['name']}")
        
        for j, tx in enumerate(transactions):
            try:
                if not isinstance(tx, dict):
                    logger.warning(f"⚠️ Транзакция {j} в {network['name']} не является словарем: {type(tx)}")
                    continue
                    
                from_addr = tx.get('from')
                to_addr = tx.get('to')
                
                if from_addr:
                    all_wallets.add(from_addr.lower())
                if to_addr:
                    all_wallets.add(to_addr.lower())
                    
            except Exception as e:
                logger.error(f"❌ Ошибка в {network['name']} транзакция {j}: {e}")
                continue
        
        total_processed += tx_count
        logger.info(f"✅ Найдено {len(transactions)} транзакций в {network['name']}")
    
    return list(all_wallets), total_processed



async def analyze_contract_single_network(contract_address, network_name, token_type, progress_callback=None):
    network = next((n for n in NETWORK_PRIORITY if n['name'] == network_name), None)
    if not network:
        return [], 0
    
    if progress_callback:
        await progress_callback(f"🔍 Анализ в {network_name}...", 0)
    
    token_contract = network['usdc_contract'] if token_type == 'usdc' else ETH_CONTRACT
    
    # ⚠️ ИСПОЛЬЗУЕМ ПОЛНУЮ ВЕРСИЮ ДЛЯ ВСЕХ СЕТЕЙ
    transactions, total_tx = await get_contract_transactions_complete(
    contract_address, network['chain_id'], token_contract if token_type == 'usdc' else None
    )
    
    logger.info(f"🔍 Начало обработки {len(transactions)} транзакций")
    
    wallets = set()
    
    for i, tx in enumerate(transactions):
        try:
            from_addr = tx.get('from', '').lower()
            to_addr = tx.get('to', '').lower()
            
            # ⚠️ ДОБАВЛЯЕМ ВСЕ ВАЛИДНЫЕ АДРЕСА
            if from_addr and len(from_addr) == 42 and from_addr.startswith('0x'):
                wallets.add(from_addr)
            if to_addr and len(to_addr) == 42 and to_addr.startswith('0x'):
                wallets.add(to_addr)
                
        except Exception as e:
            logger.error(f"❌ Ошибка обработки транзакции {i}: {e}")
            continue
    
    logger.info(f"✅ Обработано {len(transactions)} транзакций, найдено {len(wallets)} уникальных кошельков")
    
    if progress_callback:
        await progress_callback(f"✅ Анализ завершен", 100)
    
    return list(wallets), len(transactions)

async def get_allowance_via_arbitrum_rpc(wallet_address, token_contract, spender_address):
    """Проверка allowance через рабочий RPC для Arbitrum с детальным логированием"""
    try:
        # Используем твой рабочий RPC
        ARBITRUM_RPC_URL = "https://arbitrum-one-rpc.publicnode.com"
        
        # Данные для вызова функции allowance(owner, spender)
        data_payload = "0xdd62ed3e" + wallet_address[2:].zfill(64) + spender_address[2:].zfill(64)
        
        # Формируем JSON-RPC запрос
        rpc_payload = {
            "jsonrpc": "2.0",
            "method": "eth_call",
            "params": [{
                "to": token_contract,
                "data": data_payload
            }, "latest"],
            "id": 1
        }
        
        logger.info(f"🔧 [ARBITRUM_RPC_DETAILED] Calling contract...")
        logger.info(f"🔧 [ARBITRUM_RPC_DETAILED] Wallet: {wallet_address}")
        logger.info(f"🔧 [ARBITRUM_RPC_DETAILED] Token: {token_contract}")
        logger.info(f"🔧 [ARBITRUM_RPC_DETAILED] Spender: {spender_address}")
        logger.info(f"🔧 [ARBITRUM_RPC_DETAILED] Data payload: {data_payload}")
        
        await asyncio.sleep(0.5)
        
        async with aiohttp.ClientSession() as session:
            async with session.post(ARBITRUM_RPC_URL, json=rpc_payload, timeout=30) as response:
                if response.status == 200:
                    response_text = await response.text()
                    logger.info(f"🔧 [ARBITRUM_RPC_RESPONSE] Raw response: {response_text}")
                    
                    data = await response.json()
                    
                    if data.get('result'):
                        result_hex = data['result']
                        logger.info(f"🔧 [ARBITRUM_RPC_RESULT] Hex result: {result_hex}")
                        
                        if result_hex != '0x':
                            allowance_raw = int(result_hex, 16)
                            allowance_usd = allowance_raw / (10 ** 6)
                            
                            short_wallet = wallet_address[:8] + "..." + wallet_address[-6:]
                            short_spender = spender_address[:8] + "..." + spender_address[-6:]
                            
                            logger.info(f"🎯 [ARBITRUM_ALLOWANCE_FOUND] {short_wallet} -> {short_spender}: {allowance_usd:,.6f} USDC (raw: {allowance_raw})")
                            return allowance_usd
                        else:
                            logger.info(f"🔍 [ARBITRUM_EMPTY_RESULT] Empty result (0x) for {wallet_address[:8]}...")
                    else:
                        logger.info(f"🔍 [ARBITRUM_NO_RESULT] No result in response for {wallet_address[:8]}...")
                    
                    return 0
                else:
                    logger.error(f"❌ [ARBITRUM_RPC_HTTP] HTTP Error: {response.status}")
                    return 0
                    
    except Exception as e:
        logger.error(f"❌ [ARBITRUM_RPC_ERROR] {e}")
        import traceback
        logger.error(f"❌ [ARBITRUM_RPC_TRACEBACK] {traceback.format_exc()}")
        return 0

# ========== МОДУЛЬ 2: СКАНИРОВАНИЕ ALLOWANCE ==========

async def scan_allowance_single_network(wallets, token_contract, spender_address, network, progress_callback=None):
    """Оптимизированная проверка allowance - МЕНЬШЕ API CALLS"""
    results = []
    total_wallets = len(wallets)
    
    logger.info(f"🔍 [SCAN_OPTIMIZED] Сеть: {network['name']}, Кошельков: {total_wallets}")
    
    for i, wallet in enumerate(wallets):
        if progress_callback and i % 10 == 0:
            progress = int((i / total_wallets) * 100)
            await progress_callback(f"🔍 Проверка {i+1}/{total_wallets}", progress)
        
        is_eth = (token_contract == ETH_CONTRACT)
        
        # 1. Сначала проверяем баланс
        balance = await get_token_balance(wallet, token_contract, network['chain_id'], is_eth)
        
        # ЕСЛИ БАЛАНС = 0 → ПРОПУСКАЕМ проверку allowance (ЭКОНОМИЯ API CALL)
        if balance <= 0:
            continue
            
        # 2. Только если баланс > 0, проверяем allowance
        if is_eth:
            allowance = balance
        else:
            if network['chain_id'] == BASE_CHAIN_ID:
                allowance = await get_allowance_via_rpc(wallet, token_contract, spender_address)
            else:
                allowance = await get_token_allowance(wallet, token_contract, spender_address, network['chain_id'])
        
        if allowance > 0:
            results.append({
                'address': wallet,
                'allowance': allowance,
                'balance': balance,
                'available': min(allowance, balance),
                'network': network['name']
            })
            logger.info(f"🎯 [ALLOWANCE_FOUND] {wallet[:8]}...: {allowance:.6f} USDC")
        
        await asyncio.sleep(0.1)
    
    # Сортируем и добавляем ранги
    results.sort(key=lambda x: x['allowance'], reverse=True)
    for i, result in enumerate(results):
        result['rank'] = i + 1
    
    logger.info(f"✅ [SCAN_OPTIMIZED] Найдено {len(results)} кошельков, сэкономлено API calls: {total_wallets - len(results)}")
    return results

async def scan_allowance_parallel(wallets, token_contract, spender_address, network, progress_callback=None, batch_size=2):
    """ГАРАНТИРОВАННАЯ проверка ВСЕХ кошельков с retry механизмом"""
    results = []
    total_wallets = len(wallets)
    checked_wallets = 0
    skipped_due_to_zero_balance = 0
    
    logger.info(f"🔍 [GUARANTEED_SCAN] Запуск проверки {total_wallets} кошельков")
    
    async def check_single_wallet_with_retry(wallet, max_retries=3):
        """Проверка кошелька с повторными попытками при ошибках"""
        for attempt in range(max_retries):
            try:
                is_eth = (token_contract == ETH_CONTRACT)
                
                # 1. СНАЧАЛА проверяем баланс (это надежный запрос)
                balance = await get_token_balance(wallet, token_contract, network['chain_id'], is_eth)
                
                # Если баланс = 0 → логируем и пропускаем (это нормально)
                if balance <= 0:
                    return {'status': 'zero_balance', 'wallet': wallet, 'balance': balance}
                
                # 2. Если ETH → allowance = баланс
                if is_eth:
                    return {
                        'status': 'success', 
                        'wallet': wallet, 
                        'allowance': balance, 
                        'balance': balance,
                        'available': balance,
                        'network': network['name']
                    }
                
                # 3. Для USDC: проверяем allowance с паузой между попытками
                if attempt > 0:
                    retry_delay = 1.0 * attempt  # Увеличивающаяся пауза
                    logger.info(f"🔄 Повторная попытка {attempt} для {wallet[:8]}..., пауза {retry_delay}с")
                    await asyncio.sleep(retry_delay)
                
                allowance = await get_allowance_robust(wallet, token_contract, spender_address, network['chain_id'])
                
                if allowance > 0:
                    return {
                        'status': 'success',
                        'wallet': wallet,
                        'allowance': allowance,
                        'balance': balance,
                        'available': min(allowance, balance),
                        'network': network['name']
                    }
                else:
                    return {
                        'status': 'no_allowance',
                        'wallet': wallet,
                        'balance': balance,
                        'allowance': 0
                    }
                    
            except Exception as e:
                logger.warning(f"⚠️ Попытка {attempt+1} не удалась для {wallet[:8]}...: {e}")
                if attempt == max_retries - 1:  # Последняя попытка тоже failed
                    logger.error(f"❌ Все попытки не удались для {wallet}")
                    return {'status': 'error', 'wallet': wallet, 'error': str(e)}
        
        return {'status': 'error', 'wallet': wallet, 'error': 'Max retries exceeded'}
    
    # Обрабатываем кошельки МАЛЕНЬКИМИ БАТЧАМИ с БОЛЬШИМИ ПАУЗАМИ
    for batch_start in range(0, total_wallets, batch_size):
        batch_end = min(batch_start + batch_size, total_wallets)
        batch = wallets[batch_start:batch_end]
        current_batch = batch_start // batch_size + 1
        total_batches = (total_wallets + batch_size - 1) // batch_size
        
        logger.info(f"📦 Батч {current_batch}/{total_batches}: кошельки {batch_start}-{batch_end}")
        
        # Создаем задачи для текущего батча
        batch_tasks = []
        for wallet in batch:
            task = asyncio.create_task(check_single_wallet_with_retry(wallet))
            batch_tasks.append(task)
        
        # Ждем завершения ВСЕХ задач в батче
        batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
        
        # Обрабатываем результаты батча
        batch_success = 0
        batch_zero_balance = 0
        batch_no_allowance = 0
        batch_errors = 0
        
        for i, result in enumerate(batch_results):
            wallet = batch[i]
            
            if isinstance(result, Exception):
                logger.error(f"❌ Критическая ошибка для {wallet[:8]}...: {result}")
                batch_errors += 1
                continue
            
            if result['status'] == 'success':
                results.append({
                    'address': result['wallet'],
                    'allowance': result['allowance'],
                    'balance': result['balance'],
                    'available': result['available'],
                    'network': result['network']
                })
                batch_success += 1
                checked_wallets += 1
                
            elif result['status'] == 'zero_balance':
                batch_zero_balance += 1
                checked_wallets += 1
                skipped_due_to_zero_balance += 1
                
            elif result['status'] == 'no_allowance':
                batch_no_allowance += 1
                checked_wallets += 1
                
            elif result['status'] == 'error':
                batch_errors += 1
                checked_wallets += 1
        
        # Детальная статистика батча
        logger.info(f"✅ Батч {current_batch} завершен: "
                   f"Успешно={batch_success}, "
                   f"Нулевой баланс={batch_zero_balance}, "
                   f"Нет allowance={batch_no_allowance}, "
                   f"Ошибки={batch_errors}")
        
        # Обновляем прогресс
        if progress_callback:
            progress = int((batch_end / total_wallets) * 100)
            status_text = (f"🔍 Проверка {batch_end}/{total_wallets}\n"
                          f"✅ Найдено: {len(results)}\n"
                          f"💸 Нулевой баланс: {skipped_due_to_zero_balance}")
            await progress_callback(status_text, progress)
        
        # ПАУЗА между батчами для избежания rate limit
        if batch_end < total_wallets:
            pause_time = 1.0  # 1 секунда между батчами
            logger.info(f"⏳ Пауза {pause_time}с перед следующим батчем...")
            await asyncio.sleep(pause_time)
    
    # ФИНАЛЬНАЯ СТАТИСТИКА
    logger.info(f"📊 [GUARANTEED_SCAN] ИТОГО: "
               f"Кошельков={total_wallets}, "
               f"Проверено={checked_wallets}, "
               f"Найдено allowance={len(results)}, "
               f"Нулевой баланс={skipped_due_to_zero_balance}, "
               f"Ошибки={total_wallets - checked_wallets}")
    
    # Убедимся, что все кошельки обработаны
    if checked_wallets != total_wallets:
        logger.error(f"🚨 ВНИМАНИЕ: Проверено {checked_wallets} из {total_wallets} кошельков!")
    
    # Сортируем результаты
    results.sort(key=lambda x: x['allowance'], reverse=True)
    for i, result in enumerate(results):
        result['rank'] = i + 1
    
    return results

# ========== МОДУЛЬ 3: ЭКСПОРТ В CSV И EXCEL ==========

async def export_to_csv(results, filename_prefix):
    """Экспорт результатов в CSV"""
    logger.info(f"🔧 [EXPORT_DEBUG] Exporting {len(results)} results")
    
    if not results:
        logger.warning("⚠️ [EXPORT_DEBUG] No results to export")
        return None
    
    try:
        output = StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        writer.writerow(['Rank', 'Address', 'Allowance', 'Balance', 'Available', 'Network'])
        
        # Данные
        for result in results:
            writer.writerow([
                result['rank'],
                result['address'],
                f"{result['allowance']:.6f}",
                f"{result['balance']:.6f}",
                f"{result['available']:.6f}",
                result['network']
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        logger.info(f"✅ [EXPORT_DEBUG] CSV created successfully, size: {len(csv_content)} bytes")
        return BytesIO(csv_content.encode('utf-8'))
        
    except Exception as e:
        logger.error(f"❌ [EXPORT_DEBUG] Error creating CSV: {e}")
        return None

async def export_wallets_to_txt(wallets, filename_prefix):
    """Экспорт кошельков в TXT файл (каждый с новой строки)"""
    try:
        if not wallets:
            logger.warning("⚠️ [TXT_EXPORT] No wallets to export")
            return None
        
        # Создаем содержимое TXT файла
        txt_content = "\n".join(wallets)
        
        logger.info(f"✅ [TXT_EXPORT] Created TXT file with {len(wallets)} wallets")
        return BytesIO(txt_content.encode('utf-8'))
        
    except Exception as e:
        logger.error(f"❌ [TXT_EXPORT] Error creating TXT file: {e}")
        return None

async def export_to_excel(results, filename_prefix):
    """Экспорт результатов в Excel с форматированием"""
    try:
        if not results:
            logger.warning("⚠️ [EXCEL_EXPORT] No results to export")
            return None
        
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Allowance Results"
        
        # Заголовки столбцов
        headers = ['Rank', 'Address', 'Allowance (USDC)', 'Balance (USDC)', 'Available (USDC)', 'Network']
        sheet.append(headers)
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Применяем стили к заголовкам
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Добавляем данные
        for result in results:
            sheet.append([
                result['rank'],
                result['address'],
                result['allowance'],
                result['balance'],
                result['available'],
                result['network']
            ])
        
        # Форматируем числовые колонки
        number_columns = ['C', 'D', 'E']  # Allowance, Balance, Available
        for col_letter in number_columns:
            for row in range(2, len(results) + 2):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '0.000000'
        
        # Настраиваем ширину колонок
        column_widths = {
            'A': 8,   # Rank
            'B': 66,  # Address
            'C': 15,  # Allowance
            'D': 15,  # Balance
            'E': 15,  # Available
            'F': 12   # Network
        }
        
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        # Добавляем автофильтр
        sheet.auto_filter.ref = f"A1:F{len(results) + 1}"
        
        # Замораживаем заголовки
        sheet.freeze_panes = "A2"
        
        # Сохраняем в BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"✅ [EXCEL_EXPORT] Created Excel file with {len(results)} results")
        return excel_buffer
        
    except Exception as e:
        logger.error(f"❌ [EXCEL_EXPORT] Error creating Excel file: {e}")
        return None

# ========== ОБРАБОТЧИКИ КОМАНД ==========

@dp.message(Command("start"))
async def start_command(message: types.Message, state: FSMContext):
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    await state.clear()
    await message.answer(
        "👑 ПАНЕЛЬ АДМИНИСТРАТОРА - АНАЛИЗАТОР КОНТРАКТОВ\n\n"
        "📊 Доступные функции:\n"
        "• 🔍 Анализ контракта - получение списка кошельков\n"
        "• 💰 Проверка allowance - массовая проверка разрешений\n"
        "• 🛠️ Диагностика - проверка статуса API",
        reply_markup=get_admin_menu_keyboard()
    )

@dp.message(Command("test_allowance"))
async def test_allowance_command(message: types.Message):
    """Тестовая команда для проверки allowance"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_spender = "0xdf4ff5122170fe28a750d3a1d2b65bb202dd0dd2"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("🧪 Запуск теста allowance...")
    
    # Тестируем все методы
    methods = [
        ("Basescan API", get_allowance_basescan(test_wallet, test_token, test_spender)),
        ("Etherscan V2", get_token_allowance(test_wallet, test_token, test_spender, BASE_CHAIN_ID)),
        ("Direct Call", get_allowance_direct(test_wallet, test_token, test_spender, BASE_CHAIN_ID))
    ]
    
    results = []
    for method_name, method_call in methods:
        try:
            result = await method_call
            results.append(f"• {method_name}: {result:.6f} USDC")
            logger.info(f"🧪 [TEST] {method_name}: {result:.6f} USDC")
        except Exception as e:
            results.append(f"• {method_name}: ERROR - {str(e)}")
            logger.error(f"🧪 [TEST] {method_name}: ERROR - {e}")
    
    result_text = "🧪 **РЕЗУЛЬТАТЫ ТЕСТА:**\n\n" + "\n".join(results)
    await message.answer(result_text, parse_mode="Markdown")

@dp.message(lambda message: message.text == "🔍 Анализ контракта")
async def start_analysis(message: types.Message, state: FSMContext):
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    await message.answer(
        "🔍 **Собираем кошельки которые взаимодействовали с этим контрактом**\n\n"
        "Введите адрес контракта для анализа:\n\n"
        "💡 Пример: `0xdf4ff5122170fe28a750d3a1d2b65bb202dd0dd2`",
        parse_mode="Markdown"
    )
    await state.set_state(AnalyzeState.waiting_for_contract)

@dp.message(AnalyzeState.waiting_for_contract)
async def handle_contract_address(message: types.Message, state: FSMContext):
    contract_address = message.text.strip()
    
    if not (contract_address.startswith('0x') and len(contract_address) == 42):
        await message.answer("❌ Неверный формат адреса контракта. Попробуйте снова:")
        return
    
    await state.update_data(contract_address=contract_address)
    
    await message.answer(
        "🌐 Выберите режим анализа:\n\n"
        "• 🔍 Одна сеть - анализ в конкретной сети\n"
        "• 🌐 Все сети - анализ во всех сетях (Base, Arbitrum, Optimism)",
        reply_markup=get_analysis_mode_keyboard()
    )
    await state.set_state(AnalyzeState.waiting_for_mode)

async def start_allowance_check(user_id, network_choice, message=None):
    """Запускает проверку allowance с текущими параметрами"""
    try:
        user_data = current_allowance_data.get(user_id, {})
        wallets = user_data.get('wallets', [])
        token_type = user_data.get('token_type', 'usdc')
        spender_address = user_data.get('spender')
        progress_data = user_data.get('progress_data')
        
        if not wallets or not spender_address:
            if message:
                await message.answer("❌ Ошибка данных. Начните заново.")
            return
        
        token_contract = USDC_CONTRACT_BASE if token_type == 'usdc' else ETH_CONTRACT
        
        if message:
            progress_msg = await message.answer("🛡️ Запуск проверки allowance...")
        else:
            progress_msg = None
        
        async def progress_callback(text, progress):
            if progress_msg:
                try:
                    progress_bar = "█" * (progress // 10) + "░" * (10 - progress // 10)
                    await progress_msg.edit_text(f"{text}\n📊 Прогресс: {progress_bar} {progress}%")
                except Exception:
                    pass
        
        # Запускаем проверку
        if network_choice == 'all':
            all_results = []
            for network in NETWORK_PRIORITY:
                await progress_callback(f"🔍 Проверка в {network['name']}...", 0)
                network_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, user_id=user_id
                )
                all_results.extend(network_results)
        else:
            network = next((n for n in NETWORK_PRIORITY if n['name'].lower() == network_choice), None)
            if network:
                # Передаем прогресс для продолжения
                all_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, 
                    user_id=user_id
                )   
            else:
                if message:
                    await message.answer("❌ Сеть не найдена")
                return
        
        # 📊 ФИНАЛЬНЫЙ ОТЧЕТ
        if all_results:
            result_text = (
                f"😈 **ПРОВЕРКА ЗАВЕРШЕНА**\n\n"
                f"📊 Статистика:\n"
                f"• 👛 Проверено кошельков: {len(wallets):,}\n"
                f"• ✅ Найдено allowance: {len(all_results):,}\n"
                f"• 💸 Нулевой баланс: {len(wallets) - len(all_results):,}\n"
            )
            
            # Экспорт результатов
            excel_file = await export_to_excel(all_results, f"allowance_{network_choice}")
            csv_file = await export_to_csv(all_results, f"allowance_{network_choice}")
            
            if excel_file:
                await message.answer_document(
                    types.BufferedInputFile(
                        excel_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            elif csv_file:
                await message.answer_document(
                    types.BufferedInputFile(
                        csv_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            else:
                await message.answer(result_text, parse_mode="Markdown")
                
        else:
            await message.answer(
                f"❌ **ALLOWANCE НЕ НАЙДЕН**\n\n"
                f"Проверено: {len(wallets):,} кошельков\n"
                f"Spender: `{spender_address}`\n"
                f"Токен: {token_type.upper()}\n"
                f"Сеть: {network_choice.title()}"
            )
        
    except Exception as e:
        logger.error(f"❌ Allowance check error: {e}")
        if message:
            await message.answer(f"❌ Ошибка проверки: {str(e)}")    

@dp.callback_query(lambda c: c.data.startswith('mode_'))
async def handle_analysis_mode(callback: types.CallbackQuery, state: FSMContext):
    mode = callback.data.replace('mode_', '')
    await state.update_data(analysis_mode=mode)
    
    if mode == 'single':
        await callback.message.edit_text(
            "🔵 **Выберите сеть для анализа:**",
            reply_markup=get_network_selection_keyboard()
        )
        await state.set_state(AnalyzeState.waiting_for_network)
    else:
        await callback.message.edit_text(
            "💎 Выберите токен для анализа:",
            reply_markup=get_token_selection_keyboard()
        )
        await state.set_state(AnalyzeState.waiting_for_token)
    
    await callback.answer()

@dp.callback_query(lambda c: c.data.startswith('network_'))
async def handle_network_selection(callback: types.CallbackQuery, state: FSMContext):
    network_name = callback.data.replace('network_', '').title()
    await state.update_data(network=network_name)
    
    await callback.message.edit_text(
        "💎 Выберите токен для анализа:",
        reply_markup=get_token_selection_keyboard()
    )
    await state.set_state(AnalyzeState.waiting_for_token)
    await callback.answer()

# ========== ДИАГНОСТИКА ==========

@dp.message(lambda message: message.text == "🛠️ Диагностика")
async def diagnostics_command(message: types.Message):
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    diagnostic_text = "🔧 **ДИАГНОСТИКА API**\n\n"
    
    # Тестируем все сети
    for network in NETWORK_PRIORITY:
        test_params = {
            'module': 'stats',
            'action': 'ethprice'
        }
        
        data = await make_etherscan_request(network['chain_id'], test_params)
        
        if data and (data.get('status') == '1' or data.get('message') == 'OK'):
            eth_price = data.get('result', {}).get('ethusd', 'N/A')
            diagnostic_text += f"✅ {network['name']}: Работает (ETH: ${eth_price})\n"
        else:
            error_msg = data.get('message', 'Unknown error') if data else 'No response'
            diagnostic_text += f"❌ {network['name']}: {error_msg}\n"
    
    await message.answer(diagnostic_text, parse_mode="Markdown")

# ========== ОБРАБОТЧИКИ ALLOWANCE ==========

@dp.message(lambda message: message.text == "💰 Проверка allowance")
async def start_allowance_scan(message: types.Message, state: FSMContext):
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    # Очищаем предыдущее состояние
    await state.clear()
    
    # Проверяем есть ли сохраненный прогресс
    progress_data = load_progress(message.from_user.id)
    
    if progress_data:
        # Есть незавершенная сессия - предлагаем выбор
        
        # БЕЗОПАСНОЕ ПОЛУЧЕНИЕ ВРЕМЕНИ НАЧАЛА
        start_time_str = progress_data.get('start_time')
        if start_time_str:
            try:
                start_time = datetime.fromisoformat(start_time_str).strftime('%d.%m %H:%M')
            except (ValueError, TypeError):
                start_time = "неизвестно"
        else:
            start_time = "неизвестно"
        
        progress_text = (
            f"🔄 **НАЙДЕНА НЕЗАВЕРШЕННАЯ ПРОВЕРКА**\n\n"
            f"📊 Прогресс: {progress_data.get('current_index', 0)}/{progress_data.get('total_wallets', 0)} кошельков\n"
            f"✅ Найдено: {len(progress_data.get('found_allowances', []))} allowance\n"
            f"💸 Нулевых: {progress_data.get('zero_balance_count', 0)}\n"
            f"🌐 Сеть: {progress_data.get('network', 'неизвестно')}\n"
            f"🎯 Spender: `{progress_data.get('spender', '')[:10]}...`\n\n"
            f"💎 Токен: {'USDC' if progress_data.get('token_contract') != ETH_CONTRACT else 'ETH'}\n"
            f"⏰ Начато: {start_time}"
        )
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Продолжить проверку", callback_data="continue_existing_scan")],
            [InlineKeyboardButton(text="🆕 Начать новую проверку", callback_data="start_new_scan")],
            [InlineKeyboardButton(text="🗑️ Удалить прогресс", callback_data="delete_progress")]
        ])
        
        await message.answer(progress_text, reply_markup=keyboard, parse_mode="Markdown")
        
    else:
        # Нет прогресса - стандартный флоу
        await message.answer(
            "💰 ПРОВЕРКА ALLOWANCE\n\n"
            "📋 Выберите источник кошельков:",
            reply_markup=get_allowance_source_keyboard()
        )
        await state.set_state(AllowanceState.waiting_for_source)

@dp.callback_query(lambda c: c.data.startswith('source_'))
async def handle_allowance_source(callback: types.CallbackQuery, state: FSMContext):
    source = callback.data.replace('source_', '')
    
    if source == 'analysis':
        user_sessions_list = []
        
        # Ищем все сессии текущего пользователя
        for sid, data in contract_analysis_results.items():
            if sid.startswith(str(callback.from_user.id)):
                # Получаем адрес контракта из user_sessions если доступно
                contract_addr = "Unknown"
                if sid in user_sessions:
                    contract_addr = user_sessions[sid].get('contract_address', 'Unknown')
                
                wallets_count = len(data.get('wallets', []))
                # Показываем только последнюю часть адреса для компактности
                short_addr = contract_addr[:10] + "..." + contract_addr[-8:] if contract_addr != "Unknown" else "Unknown"
                user_sessions_list.append(f"`{sid}`: {short_addr} ({wallets_count} кошельков)")
        
        if not user_sessions_list:
            await callback.message.edit_text(
                "❌ Нет сохраненных результатов анализа.\n"
                "Сначала выполните анализ контракта."
            )
            return
        
        sessions_text = "\n".join(user_sessions_list[:5])
        await callback.message.edit_text(
            f"📋 **Доступные результаты анализа:**\n\n{sessions_text}\n\n"
            f"Введите ID сессии:",
            parse_mode="Markdown"
        )
        await state.set_state(AllowanceState.waiting_for_wallets)
        
    else:
        await callback.message.edit_text(
            "📁 Загрузка кошельков...\n\n"
            "Отправьте TXT файл с кошельками (каждый с новой строки):\n\n"
            "Или CSV файл выгрузкой из scan "
        )
        await state.set_state(AllowanceState.waiting_for_wallets)
    
    await callback.answer()

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ ПАРСИНГА ФАЙЛОВ ==========

def is_valid_wallet_address(address):
    """Проверяет, является ли строка валидным адресом кошелька"""
    if not address or not isinstance(address, str):
        return False
    
    address_clean = address.strip()
    
    return (address_clean.startswith('0x') and 
            len(address_clean) == 42 and 
            all(c in '0123456789abcdefABCDEF' for c in address_clean[2:]))

def extract_wallet_addresses_from_csv(content):
    """Извлекает адреса кошельков из CSV контента с транзакциями"""
    wallets = set()
    original_rows = 0
    
    try:
        # Считаем общее количество строк
        lines = content.split('\n')
        original_rows = len([line for line in lines if line.strip()])
        
        # Пробуем разные разделители
        for delimiter in [',', ';', '\t']:
            csv_reader = csv.reader(StringIO(content), delimiter=delimiter)
            for row_num, row in enumerate(csv_reader):
                if len(row) >= 6:  # Нужны как минимум 6 столбцов
                    # Столбец From (индекс 4)
                    if len(row) > 4:
                        from_addr = str(row[4]).strip().strip('"').strip("'")
                        if is_valid_wallet_address(from_addr):
                            wallets.add(from_addr.lower())
                    
                    # Столбец To (индекс 5)
                    if len(row) > 5:
                        to_addr = str(row[5]).strip().strip('"').strip("'")
                        if is_valid_wallet_address(to_addr):
                            wallets.add(to_addr.lower())
        
        # Если не нашли через CSV reader, пробуем найти адреса регулярными выражениями
        if not wallets:
            import re
            address_pattern = re.compile(r'0x[a-fA-F0-9]{40}')
            matches = address_pattern.findall(content)
            wallets.update([match.lower() for match in matches])
        
    except Exception as e:
        logger.error(f"❌ Error parsing CSV: {e}")
    
    return list(wallets), original_rows

def parse_wallets_from_content(content, file_name=None):
    """Парсит кошельки из содержимого файла"""
    wallets = []
    
    try:
        if file_name and file_name.lower().endswith('.csv'):
            # CSV файл - парсим транзакции и извлекаем адреса From и To
            csv_reader = csv.reader(StringIO(content))
            
            for row in csv_reader:
                if len(row) >= 6:  # Проверяем, что есть достаточно столбцов
                    # 5-й столбец - From (индекс 4)
                    from_address = row[4].strip().strip('"').strip("'")
                    # 6-й столбец - To (индекс 5)  
                    to_address = row[5].strip().strip('"').strip("'")
                    
                    if is_valid_wallet_address(from_address):
                        wallets.append(from_address.lower())
                    if is_valid_wallet_address(to_address):
                        wallets.append(to_address.lower())
        else:
            # TXT файл или другой формат - каждый кошелек на новой строке
            lines = content.split('\n')
            for line in lines:
                line_clean = line.strip()
                if is_valid_wallet_address(line_clean):
                    wallets.append(line_clean.lower())
        
        # Убираем дубликаты и пустые адреса
        wallets = list(set([w for w in wallets if w]))
        
        logger.info(f"✅ [PARSE_WALLETS] Parsed {len(wallets)} wallets from {file_name or 'content'}")
        
    except Exception as e:
        logger.error(f"❌ Error parsing wallets from content: {e}")
    
    return wallets   

# ========== ОБРАБОТЧИКИ ALLOWANCE ==========

@dp.message(AllowanceState.waiting_for_wallets)
async def handle_wallets_input(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        
        if message.document:
            file_info = await bot.get_file(message.document.file_id)
            downloaded_file = await bot.download_file(file_info.file_path)
            
            content = downloaded_file.read().decode('utf-8')
            
            # Определяем тип файла по расширению
            file_extension = message.document.file_name.lower().split('.')[-1] if message.document.file_name else ''
            
            wallets = []
            
            if file_extension == 'txt' or ('\n' in content and ',' not in content):
                # TXT файл - каждый кошелек на новой строке
                wallets = [line.strip() for line in content.split('\n') if line.strip()]
                
            elif file_extension == 'csv' or (',' in content and '\n' in content):
                # CSV файл - парсим транзакции и извлекаем адреса From и To
                csv_reader = csv.reader(StringIO(content))
                
                for row in csv_reader:
                    if len(row) >= 6:  # Проверяем, что есть достаточно столбцов
                        # 5-й столбец - From (индекс 4)
                        from_address = row[4].strip().strip('"').strip("'")
                        # 6-й столбец - To (индекс 5)  
                        to_address = row[5].strip().strip('"').strip("'")
                        
                        # Проверяем валидность адреса From
                        if (from_address.startswith('0x') and 
                            len(from_address) == 42 and 
                            all(c in '0123456789abcdefABCDEF' for c in from_address[2:])):
                            wallets.append(from_address.lower())
                        
                        # Проверяем валидность адреса To
                        if (to_address.startswith('0x') and 
                            len(to_address) == 42 and 
                            all(c in '0123456789abcdefABCDEF' for c in to_address[2:])):
                            wallets.append(to_address.lower())
            
            else:
                # Пробуем общий парсинг для других форматов
                lines = content.split('\n')
                for line in lines:
                    line_clean = line.strip()
                    if (line_clean.startswith('0x') and 
                        len(line_clean) == 42 and 
                        all(c in '0123456789abcdefABCDEF' for c in line_clean[2:])):
                        wallets.append(line_clean.lower())
            
            # Убираем дубликаты и пустые адреса
            wallets = list(set([w for w in wallets if w]))
            
            if not wallets:
                await message.answer(
                    "❌ Не удалось найти валидные адреса кошельков в файле.\n\n"
                    "Для CSV файлов убедитесь, что файл содержит столбцы с адресами в формате 0x... "
                    "в 5-м и 6-м столбцах (From и To)."
                )
                return
            
            # Получаем текущие данные пользователя
            user_data = current_allowance_data.get(user_id, {})
            
            # Проверяем, это продолжение существующей проверки?
            is_continuation = user_data.get('source') == 'progress'
            
            if is_continuation:
                # ПРОДОЛЖЕНИЕ СУЩЕСТВУЮЩЕЙ ПРОВЕРКИ
                progress_data = user_data.get('progress_data', {})
                expected_wallet_count = progress_data.get('total_wallets', 0)
                
                if len(wallets) != expected_wallet_count:
                    await message.answer(
                        f"❌ **НЕСОВПАДЕНИЕ КОЛИЧЕСТВА КОШЕЛЬКОВ**\n\n"
                        f"Ожидалось: {expected_wallet_count} кошельков\n"
                        f"Загружено: {len(wallets)} кошельков\n\n"
                        f"💡 Для продолжения загрузите исходный файл с {expected_wallet_count} кошельками",
                        parse_mode="Markdown"
                    )
                    return
                
                # Сохраняем кошельки и продолжаем
                current_allowance_data[user_id]['wallets'] = wallets
                
                # Пропускаем выбор токена и спендера - берем из прогресса
                token_type = user_data.get('token_type', 'usdc')
                spender_address = user_data.get('spender')
                network_from_progress = user_data.get('network_from_progress', 'Base')
                current_index = progress_data.get('current_index', 0)
                
                await message.answer(
                    f"✅ **ПРОВЕРКА ВОССТАНОВЛЕНА**\n\n"
                    f"🔄 Продолжаем с {current_index}-го кошелька\n"
                    f"📊 Всего кошельков: {len(wallets)}\n\n"
                    f"💎 Токен: {token_type.upper()}\n"
                    f"🎯 Spender: `{spender_address}`\n"
                    f"🌐 Сеть: {network_from_progress}",
                    parse_mode="Markdown"
                )
                
                # Автоматически запускаем проверку в той же сети
                network_choice = network_from_progress.lower()
                await start_allowance_check(user_id, network_choice, message)
                return
                
            else:
                # СТАНДАРТНАЯ ОБРАБОТКА (новая проверка)
                # Сохраняем в глобальную переменную
                current_allowance_data[user_id] = {
                    'wallets': wallets,
                    'source': 'file',
                    'file_type': file_extension
                }
                
                logger.info(f"✅ Загружено {len(wallets)} уникальных кошельков из {file_extension.upper()} файла")
                
                await message.answer(
                    f"💎 Токен для проверки allowance:\n\n"
                    f"👛 Загружено кошельков: {len(wallets):,}",
                    reply_markup=get_token_selection_keyboard()
                )
                await state.set_state(AllowanceState.waiting_for_token_allowance)
                
        elif message.text:
            # Обработка ID сессии (существующий код)
            session_id = message.text.strip()
            if session_id in contract_analysis_results:
                data = contract_analysis_results[session_id]
                wallets = data.get('wallets', [])
                if wallets:
                    # Сохраняем в глобальную переменную
                    current_allowance_data[user_id] = {
                        'wallets': wallets,
                        'source': 'analysis',
                        'session_id': session_id
                    }
                    logger.info(f"✅ [ALLOWANCE] Loaded {len(wallets)} wallets for user {user_id}")
                    
                    await message.answer(
                        f"💎 **Выберите токен для проверки allowance:**\n\n"
                        f"👛 Загружено кошельков: {len(wallets):,}",
                        reply_markup=get_token_selection_keyboard()
                    )
                    await state.set_state(AllowanceState.waiting_for_token_allowance)
                else:
                    await message.answer("❌ В этой сессии нет кошельков. Попробуйте другую:")
                    return
            else:
                await message.answer("❌ Сессия не найдена. Попробуйте снова:")
                return
        
        else:
            await message.answer("❌ Отправьте TXT/CSV файл или ID сессии:")
            return
        
    except Exception as e:
        logger.error(f"❌ Error in handle_wallets_input: {e}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        await message.answer("❌ Произошла ошибка при загрузке кошельков. Убедитесь, что файл имеет правильный формат.")

@dp.callback_query(AllowanceState.waiting_for_token_allowance)
async def handle_allowance_token_selection(callback: types.CallbackQuery, state: FSMContext):
    try:
        user_id = callback.from_user.id
        token_type = callback.data.replace('token_', '')
        
        logger.info(f"🔧 [ALLOWANCE_TOKEN] START - User: {user_id}, Token: {token_type}")
        
        # Проверяем глобальные данные
        logger.info(f"🔧 [ALLOWANCE_TOKEN] Global data keys: {list(current_allowance_data.keys())}")
        
        user_data = current_allowance_data.get(user_id, {})
        wallets = user_data.get('wallets', [])
        session_id = user_data.get('session_id')
        
        logger.info(f"🔧 [ALLOWANCE_TOKEN] User data: {user_data}, Wallets count: {len(wallets)}")
        
        # Если wallets пустые, пытаемся восстановить
        if not wallets:
            logger.warning("⚠️ [ALLOWANCE_TOKEN] No wallets in global data, attempting recovery...")
            
            # Способ 1: Из состояния FSM
            state_data = await state.get_data()
            logger.info(f"🔧 [ALLOWANCE_TOKEN] State data: {state_data}")
            
            # Способ 2: Из последней сессии пользователя
            user_sessions_list = [
                sid for sid in contract_analysis_results.keys() 
                if sid.startswith(str(user_id))
            ]
            logger.info(f"🔧 [ALLOWANCE_TOKEN] User sessions: {user_sessions_list}")
            
            if user_sessions_list:
                last_session = user_sessions_list[-1]
                session_data = contract_analysis_results[last_session]
                wallets = session_data.get('wallets', [])
                session_id = last_session
                
                if wallets:
                    logger.info(f"✅ [ALLOWANCE_TOKEN] Recovered {len(wallets)} wallets from last session: {last_session}")
                    # Сохраняем в глобальные данные
                    current_allowance_data[user_id] = {
                        'wallets': wallets,
                        'source': 'analysis', 
                        'session_id': session_id,
                        'token_type': token_type
                    }
        
        # Если все еще нет wallets - ошибка
        if not wallets:
            logger.error(f"❌ [ALLOWANCE_TOKEN] No wallets found after all recovery attempts")
            await callback.answer("❌ Не удалось найти кошельки для проверки. Начните заново.", show_alert=True)
            return
        
        # Обновляем глобальные данные
        current_allowance_data[user_id]['token_type'] = token_type
        
        logger.info(f"✅ [ALLOWANCE_TOKEN] SUCCESS - User {user_id}: {len(wallets)} wallets, token: {token_type}")
        
        await callback.message.edit_text(
            f"🔍 **Введите адрес spender для {token_type.upper()}:**\n\n"
            f"👛 Кошельков для проверки: {len(wallets):,}\n\n"
            "📝 Введите адрес:",
            parse_mode="Markdown"
        )
        await state.set_state(AllowanceState.waiting_for_spender)
        await callback.answer()
        
    except Exception as e:
        logger.error(f"❌ [ALLOWANCE_TOKEN] CRITICAL ERROR: {e}")
        import traceback
        logger.error(f"❌ [ALLOWANCE_TOKEN] Traceback: {traceback.format_exc()}")
        await callback.answer(f"❌ Критическая ошибка: {str(e)}", show_alert=True)

@dp.message(AllowanceState.waiting_for_spender)
async def handle_spender_address(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        spender_address = message.text.strip()
        
        if not (spender_address.startswith('0x') and len(spender_address) == 42):
            await message.answer("❌ Неверный формат адреса. Попробуйте снова:")
            return
        
        # Получаем текущие данные пользователя
        user_data = current_allowance_data.get(user_id, {})
        
        # Обновляем данные
        user_data['spender'] = spender_address
        current_allowance_data[user_id] = user_data
        
        logger.info(f"✅ [SPENDER] User {user_id} set spender: {spender_address}")
        logger.info(f"✅ [SPENDER] User data keys: {list(user_data.keys())}")
        
        await message.answer(
            "🌐 Выберите сеть для проверки allowance:\n\n"
            "• 🔵 Base\n"
            "• 🔷 Arbitrum  \n"
            "• 🟠 Optimism\n"
            "• 🌐 Все сети (может занять время)",
            reply_markup=get_network_selection_keyboard_allowance()
        )
        
    except Exception as e:
        logger.error(f"❌ Error in handle_spender_address: {e}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        await message.answer("❌ Произошла ошибка. Попробуйте снова.")

@dp.callback_query(lambda c: c.data.startswith('allowance_network_'))
async def handle_allowance_network_selection(callback: types.CallbackQuery, state: FSMContext):
    try:
        user_id = callback.from_user.id
        network_choice = callback.data.replace('allowance_network_', '')
        
        # Сразу отвечаем на callback чтобы избежать "query is too old"
        await callback.answer("🔄 Запуск проверки...")
        
        # Получаем данные
        user_data = current_allowance_data.get(user_id, {})
        wallets = user_data.get('wallets', [])
        token_type = user_data.get('token_type', 'usdc')
        spender_address = user_data.get('spender')
        
        if not wallets or not spender_address:
            await callback.message.answer("❌ Ошибка данных. Начните заново.")
            return
        
        token_contract = USDC_CONTRACT_BASE if token_type == 'usdc' else ETH_CONTRACT
        
        # Отправляем новое сообщение вместо редактирования старого
        progress_msg = await callback.message.answer("🛡️ Запуск надежной проверки allowance...")
        
        async def progress_callback(text, progress):
            try:
                progress_bar = "█" * (progress // 10) + "░" * (10 - progress // 10)
                await progress_msg.edit_text(f"{text}\n📊 Прогресс: {progress_bar} {progress}%")
            except Exception as e:
                # Если сообщение устарело, создаем новое
                if "message is not modified" in str(e):
                    pass
        
        # Запускаем проверку
        if network_choice == 'all':
            all_results = []
            for network in NETWORK_PRIORITY:
                await progress_callback(f"🔍 Проверка в {network['name']}...", 0)
                network_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, user_id=callback.from_user.id
                )
                all_results.extend(network_results)
        else:
            network = next((n for n in NETWORK_PRIORITY if n['name'].lower() == network_choice), None)
            if network:
                all_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, user_id=callback.from_user.id
                )   
            else:
                await callback.message.answer("❌ Сеть не найдена")
                return
        
        # 📊 ФИНАЛЬНЫЙ ОТЧЕТ
        if all_results:
            result_text = (
                f"😈 **ПРОВЕРКА ЗАВЕРШЕНА**\n\n"
                f"📊 Статистика:\n"
                f"• 👛 Проверено кошельков: {len(wallets):,}\n"
                f"• ✅ Найдено allowance: {len(all_results):,}\n"
                f"• 💸 Нулевой баланс: {len(wallets) - len(all_results):,}\n"
            )
            
            # Экспорт результатов
            excel_file = await export_to_excel(all_results, f"allowance_{network_choice}")
            csv_file = await export_to_csv(all_results, f"allowance_{network_choice}")
            
            if excel_file:
                await callback.message.answer_document(
                    types.BufferedInputFile(
                        excel_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            elif csv_file:
                await callback.message.answer_document(
                    types.BufferedInputFile(
                        csv_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            else:
                await callback.message.answer(result_text, parse_mode="Markdown")
                
        else:
            await callback.message.answer(
                f"❌ **ALLOWANCE НЕ НАЙДЕН**\n\n"
                f"Проверено: {len(wallets):,} кошельков\n"
                f"Spender: `{spender_address}`\n"
                f"Токен: {token_type.upper()}\n"
                f"Сеть: {network_choice.title()}"
            )
        
    except Exception as e:
        logger.error(f"❌ Allowance scan error: {e}")
        await callback.message.answer(f"❌ Ошибка проверки: {str(e)}")

@dp.callback_query(lambda c: c.data in ["continue_existing_scan", "start_new_scan", "delete_progress"])
async def handle_continue_scan_choice(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    
    if callback.data == "continue_existing_scan":
        # Продолжаем существующую проверку
        progress_data = load_progress(user_id)
        
        if not progress_data:
            await callback.answer("❌ Прогресс не найден", show_alert=True)
            return
        
        # Сохраняем данные прогресса в глобальную переменную
        current_allowance_data[user_id] = {
            'wallets': [],  # Кошельки нужно будет загрузить заново
            'source': 'progress',
            'token_type': 'usdc' if progress_data['token_contract'] != ETH_CONTRACT else 'eth',
            'spender': progress_data['spender'],
            'progress_data': progress_data,
            'network_from_progress': progress_data['network']
        }
        
        await callback.message.edit_text(
            f"🔄 **ПРОДОЛЖЕНИЕ ПРОВЕРКИ**\n\n"
            f"Загрузите исходный файл с {progress_data['total_wallets']} кошельками\n"
            f"для продолжения с {progress_data['current_index']}-го кошелька\n\n"
            f"💡 **Параметры:**\n"
            f"• Сеть: {progress_data['network']}\n"
            f"• Токен: {'USDC' if progress_data['token_contract'] != ETH_CONTRACT else 'ETH'}\n"
            f"• Spender: `{progress_data['spender']}`",
            parse_mode="Markdown"
        )
        await state.set_state(AllowanceState.waiting_for_wallets)
        
    elif callback.data == "start_new_scan":
        # Начинаем новую проверку
        await callback.message.edit_text(
            "💰 ПРОВЕРКА ALLOWANCE\n\n"
            "📋 Выберите источник кошельков:",
            reply_markup=get_allowance_source_keyboard()
        )
        await state.set_state(AllowanceState.waiting_for_source)
        
    elif callback.data == "delete_progress":
        # Удаляем прогресс
        delete_progress(user_id)
        await callback.message.edit_text(
            "🗑️ **ПРОГРЕСС УДАЛЕН**\n\n"
            "Теперь вы можете начать новую проверку:",
            reply_markup=get_allowance_source_keyboard()
        )
        await state.set_state(AllowanceState.waiting_for_source)
    
    await callback.answer()

@dp.callback_query()
async def debug_all_callbacks(callback: types.CallbackQuery, state: FSMContext):
    """Временный обработчик для отладки всех callback"""
    logger.info(f"🔧 [DEBUG_CALLBACK] Data: {callback.data}, State: {await state.get_state()}")
    user_data = await state.get_data()
    logger.info(f"🔧 [DEBUG_CALLBACK] User data: {user_data}")
    
    # Если это токен для анализа контракта
    if callback.data.startswith('token_') and await state.get_state() == AnalyzeState.waiting_for_token:
        await handle_analysis_token_selection(callback, state)
    # Если это токен для allowance
    elif callback.data.startswith('token_') and await state.get_state() == AllowanceState.waiting_for_token_allowance:
        await handle_allowance_token_selection(callback, state)
    else:
        await callback.answer("❌ Неизвестная команда")

# Добавь этот временный обработчик для анализа
async def handle_analysis_token_selection(callback: types.CallbackQuery, state: FSMContext):
    """Временный обработчик для выбора токена при анализе"""
    try:
        logger.info(f"🔧 [ANALYSIS_TOKEN] Starting analysis for token: {callback.data}")
        
        # Получаем данные из состояния
        user_data = await state.get_data()
        
        # Проверяем наличие обязательных полей
        if 'contract_address' not in user_data:
            await callback.answer("❌ Ошибка: адрес контракта не найден. Начните заново.", show_alert=True)
            await state.clear()
            return
            
        contract_address = user_data['contract_address']
        analysis_mode = user_data.get('analysis_mode', 'single')
        network_name = user_data.get('network', 'Base')
        
        token_type = callback.data.replace('token_', '')
        
        session_id = f"{callback.from_user.id}_{int(time.time())}"
        
        # Сохраняем в user_sessions с полной информацией
        user_sessions[session_id] = {
            'contract_address': contract_address,
            'token_type': token_type,
            'analysis_mode': analysis_mode,
            'network': network_name
        }
        
        # Сохраняем последнюю сессию пользователя
        last_user_sessions[callback.from_user.id] = session_id
        logger.info(f"💾 [SESSION_DEBUG] Saved last session for user {callback.from_user.id}: {session_id}")
        
        progress_msg = await callback.message.edit_text("🔄 Начинаем анализ контракта...")
        
        async def progress_callback(text, progress):
            try:
                progress_bar = "█" * (progress // 10) + "░" * (10 - progress // 10)
                await progress_msg.edit_text(
                    f"🔄 {text}\n\n"
                    f"📊 Прогресс: {progress_bar} {progress}%"
                )
            except:
                pass
        
        try:
            if analysis_mode == 'all':
                wallets, total_tx = await analyze_contract_all_networks(
                    contract_address, token_type, progress_callback
                )
            else:
                wallets, total_tx = await analyze_contract_single_network(
                    contract_address, network_name, token_type, progress_callback
                )
            
            contract_analysis_results[session_id] = {
                'wallets': wallets,
                'total_transactions': total_tx,
                'timestamp': datetime.now().isoformat()
            }
            
            result_text = (
                f"✅ **АНАЛИЗ ЗАВЕРШЕН**\n\n"
                f"📄 Контракт: `{contract_address}`\n"
                f"🌐 Режим: {'Все сети' if analysis_mode == 'all' else network_name}\n"
                f"💎 Токен: {token_type.upper()}\n"
                f"👛 Найдено кошельков: {len(wallets):,}\n"
                f"📊 Обработано транзакций: {total_tx:,}\n\n"
                f"💡 Результат сохранен для проверки allowance"
            )
            
            await progress_msg.edit_text(result_text, parse_mode="Markdown")
            
            # Предлагаем скачать кошельки в TXT формате
            if wallets:
                txt_file = await export_wallets_to_txt(wallets, f"wallets_{session_id}")
                if txt_file:
                    await callback.message.answer_document(
                        types.BufferedInputFile(
                            txt_file.getvalue(),
                            filename=f"wallets_{network_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
                        ),
                        caption="📝 Список кошельков (каждый с новой строки)"
                    )
                else:
                    await callback.message.answer("❌ Не удалось создать файл с кошельками")
            else:
                await callback.message.answer("ℹ️ Кошельки не найдены для экспорта")
            
        except Exception as e:
            logger.error(f"Analysis error: {e}")
            await progress_msg.edit_text(f"❌ Ошибка анализа: {str(e)}")
        
        await state.clear()
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Token selection error: {e}")
        await callback.answer("❌ Произошла ошибка. Попробуйте снова.", show_alert=True)
        await state.clear()

# ========== ДИАГНОСТИЧЕСКИЕ КОМАНДЫ ==========

@dp.message(Command("deep_debug"))
async def deep_debug_command(message: types.Message):
    """Глубокая диагностика с проверкой через разные endpoints"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_spender = "0xdf4ff5122170fe28a750d3a1d2b65bb202dd0dd2"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("🔬 **ГЛУБОКАЯ ДИАГНОСТИКА**\n\nЗапуск всех методов проверки...")
    
    debug_results = []
    
    # Метод 1: Basescan API напрямую
    try:
        params = {
            'module': 'account',
            'action': 'tokenallowance',
            'contractaddress': test_token,
            'address': test_wallet,
            'spender': test_spender,
            'apikey': BASESCAN_API_KEY
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.get(BASESCAN_API_URL, params=params, timeout=30) as response:
                response_text = await response.text()
                debug_results.append(f"🔹 **Basescan Direct**:\nСтатус: {response.status}\nОтвет: {response_text}")
                
    except Exception as e:
        debug_results.append(f"🔹 **Basescan Direct**: Ошибка - {str(e)}")
    
    # Метод 2: Etherscan V2 API
    try:
        params = {
            'module': 'account',
            'action': 'tokenallowance',
            'contractaddress': test_token,
            'address': test_wallet,
            'spender': test_spender,
            'chainid': BASE_CHAIN_ID,
            'apikey': ETHERSCAN_API_KEY
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.get(ETHERSCAN_V2_API_URL, params=params, timeout=30) as response:
                response_text = await response.text()
                debug_results.append(f"🔹 **Etherscan V2**:\nСтатус: {response.status}\nОтвет: {response_text}")
                
    except Exception as e:
        debug_results.append(f"🔹 **Etherscan V2**: Ошибка - {str(e)}")
    
    # Метод 3: Прямой RPC вызов (имитация web3)
    try:
        # Данные для вызова функции allowance(owner, spender)
        data_payload = "0xdd62ed3e" + test_wallet[2:].zfill(64) + test_spender[2:].zfill(64)
        
        params = {
            'module': 'proxy',
            'action': 'eth_call',
            'to': test_token,
            'data': data_payload,
            'tag': 'latest',
            'apikey': BASESCAN_API_KEY
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.get(BASESCAN_API_URL, params=params, timeout=30) as response:
                response_text = await response.text()
                debug_results.append(f"🔹 **Direct Contract Call**:\nСтатус: {response.status}\nОтвет: {response_text}")
                
    except Exception as e:
        debug_results.append(f"🔹 **Direct Contract Call**: Ошибка - {str(e)}")
    
    # Метод 4: Проверка через альтернативный endpoint
    try:
        params = {
            'module': 'account',
            'action': 'tokenallowance',
            'contractaddress': test_token,
            'address': test_wallet,
            'spender': test_spender,
            'apikey': ETHERSCAN_API_KEY
        }
        
        # Пробуем старый endpoint
        alt_url = "https://api.basescan.org/api"
        async with aiohttp.ClientSession() as session:
            async with session.get(alt_url, params=params, timeout=30) as response:
                response_text = await response.text()
                debug_results.append(f"🔹 **Alternative Basescan**:\nСтатус: {response.status}\nОтвет: {response_text}")
                
    except Exception as e:
        debug_results.append(f"🔹 **Alternative Basescan**: Ошибка - {str(e)}")
    
    # Формируем итоговый отчет
    result_text = "🔬 **РЕЗУЛЬТАТЫ ГЛУБОКОЙ ДИАГНОСТИКИ:**\n\n"
    result_text += "\n\n".join(debug_results)
    
    # Обрезаем если слишком длинный
    if len(result_text) > 4000:
        result_text = result_text[:4000] + "\n\n... (сообщение обрезано)"
    
    await message.answer(f"```\n{result_text}\n```", parse_mode="Markdown")

@dp.message(Command("check_balance"))
async def check_balance_command(message: types.Message):
    """Проверка только баланса"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("💰 **ПРОВЕРКА БАЛАНСА**\n\nЗапрос...")
    
    try:
        params = {
            'module': 'account',
            'action': 'tokenbalance',
            'contractaddress': test_token,
            'address': test_wallet,
            'tag': 'latest',
            'apikey': BASESCAN_API_KEY
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.get(BASESCAN_API_URL, params=params, timeout=30) as response:
                response_text = await response.text()
                data = json.loads(response_text)
                
                if data.get('status') == '1':
                    balance_raw = data.get('result', '0')
                    balance = int(balance_raw) / (10 ** 6)
                    result_text = f"✅ **БАЛАНС НАЙДЕН:**\n\n💰 {balance:.6f} USDC\n\n📊 Raw: {balance_raw}"
                else:
                    result_text = f"❌ **ОШИБКА:**\n\n{data}"
                    
                await message.answer(f"```\n{result_text}\n```", parse_mode="Markdown")
                
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.message(Command("test_v2"))
async def test_v2_command(message: types.Message):
    """Тест V2 API"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("🧪 **ТЕСТ V2 API**\n\nПроверка баланса через V2...")
    
    try:
        params = {
            'module': 'account',
            'action': 'tokenbalance',
            'contractaddress': test_token,
            'address': test_wallet,
            'tag': 'latest'
        }
        
        data = await make_etherscan_request(BASE_CHAIN_ID, params)
        
        if data:
            result_text = f"📊 **V2 API ОТВЕТ:**\n\n```json\n{json.dumps(data, indent=2)}\n```"
            
            if data.get('status') == '1' or data.get('message') == 'OK':
                balance_raw = data.get('result', '0')
                balance = int(balance_raw) / (10 ** 6)
                result_text += f"\n\n✅ **БАЛАНС:** {balance:.6f} USDC"
            else:
                result_text += f"\n\n❌ **ОШИБКА:** {data.get('message')}"
        else:
            result_text = "❌ **НЕТ ОТВЕТА ОТ API**"
            
        await message.answer(result_text, parse_mode="Markdown")
        
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.message(Command("check_max_allowance"))
async def check_max_allowance_command(message: types.Message):
    """Проверка MAX_UINT256 allowance"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_spender = "0xdf4ff5122170fe28a750d3a1d2b65bb202dd0dd2"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("🔍 **ПРОВЕРКА MAX ALLOWANCE**\n\nАнализируем данные...")
    
    MAX_UINT256 = 115792089237316195423570985008687907853269984665640564039457584007913129639935
    
    try:
        # Метод 1: Прямой вызов контракта
        data_payload = "0xdd62ed3e" + test_wallet[2:].zfill(64) + test_spender[2:].zfill(64)
        
        params = {
            'module': 'proxy',
            'action': 'eth_call',
            'to': test_token,
            'data': data_payload,
            'tag': 'latest'
        }
        
        data = await make_etherscan_request(BASE_CHAIN_ID, params)
        
        if data and data.get('result'):
            result_hex = data['result']
            
            if result_hex != '0x':
                allowance_raw = int(result_hex, 16)
                allowance_usd = allowance_raw / (10 ** 6)
                
                result_text = f"🔹 **Прямой вызов контракта**:\n"
                result_text += f"Hex результат: `{result_hex}`\n"
                result_text += f"Числовое значение: `{allowance_raw}`\n"
                result_text += f"USDC значение: `{allowance_usd:.6f}`\n\n"
                
                # Проверяем является ли это MAX_UINT256
                if allowance_raw == MAX_UINT256:
                    result_text += f"🎯 **ЭТО MAX_UINT256!**\n"
                    result_text += f"✅ **БЕСКОНЕЧНЫЙ ALLOWANCE**\n\n"
                    result_text += f"💡 Это означает, что spender может использовать ВСЕ USDC этого кошелька\n"
                elif allowance_raw > 0:
                    result_text += f"✅ **ALLOWANCE НАЙДЕН**: {allowance_usd:,.2f} USDC\n"
                else:
                    result_text += f"❌ **ALLOWANCE НЕ НАЙДЕН**\n"
            else:
                result_text = f"❌ **Пустой результат от контракта**: {result_hex}"
        else:
            error_msg = data.get('message', 'No data') if data else 'No response'
            result_text = f"❌ **Ошибка API**: {error_msg}"
            
        await message.answer(result_text, parse_mode="Markdown")
        
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")

@dp.message(Command("test_rpc"))
async def test_rpc_command(message: types.Message):
    """Тест RPC метода"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    test_wallet = "0x0bD785D8ab76dE1595e595DBc12eF337cc49Ed72"
    test_spender = "0xdf4ff5122170fe28a750d3a1d2b65bb202dd0dd2"
    test_token = "0x833589fCD6eDb6E08f4c7c32D4f71b54bdA02913"
    
    await message.answer("🔧 **ТЕСТ RPC МЕТОДА**\n\nПроверка allowance через Base RPC...")
    
    try:
        result = await get_allowance_via_rpc(test_wallet, test_token, test_spender)
        
        if result > 0:
            result_text = f"✅ **RPC ALLOWANCE НАЙДЕН:** {result:.6f} USDC\n\n"
            result_text += f"💡 Кошелек: `{test_wallet}`\n"
            result_text += f"💡 Spender: `{test_spender}`\n"
            result_text += f"💡 Токен: `{test_token}`"
        else:
            result_text = f"❌ **RPC ALLOWANCE НЕ НАЙДЕН**\n\n"
            result_text += f"Проверьте правильность адресов и наличие allowance"
            
        await message.answer(result_text, parse_mode="Markdown")
        
    except Exception as e:
        await message.answer(f"❌ Ошибка RPC: {str(e)}")

async def get_allowance_reliable(wallet_address, token_contract, spender_address, chain_id, max_retries=3):
    """Самый надежный метод проверки allowance - только через RPC"""
    for attempt in range(max_retries):
        try:
            # Пауза между попытками
            if attempt > 0:
                wait_time = attempt * 1.0
                logger.info(f"🔄 Повторная попытка {attempt} для {wallet_address[:8]}..., пауза {wait_time}с")
                await asyncio.sleep(wait_time)
            
            # Используем RPC вызов для всех сетей
            if chain_id == BASE_CHAIN_ID:
                # Для Base используем Base RPC
                allowance = await get_allowance_via_rpc(wallet_address, token_contract, spender_address)
            else:
                # Для других сетей используем прямое обращение к контракту через Etherscan proxy
                allowance = await get_allowance_direct(wallet_address, token_contract, spender_address, chain_id)
            
            if allowance > 0:
                logger.info(f"✅ [ALLOWANCE_FOUND] {wallet_address[:8]}...: {allowance:.6f} USDC")
                return allowance
            else:
                logger.info(f"ℹ️ [NO_ALLOWANCE] {wallet_address[:8]}...: 0 USDC")
                return 0
                
        except Exception as e:
            logger.warning(f"⚠️ Ошибка проверки {wallet_address[:8]}... (попытка {attempt+1}): {e}")
            continue
    
    logger.error(f"❌ Все попытки не удались для {wallet_address[:8]}...")
    return 0

async def scan_allowance_reliable(wallets, token_contract, spender_address, network, progress_callback=None, user_id=None, progress_data=None):
    """Надежная проверка ВСЕХ кошельков с кэшированием прогресса и оптимизированной скоростью"""
    
    # Создаем ID сессии
    session_id = f"{user_id}_{int(time.time())}" if user_id else f"session_{int(time.time())}"
    total_wallets = len(wallets)
    
    # Проверяем есть ли сохраненный прогресс ИЛИ переданные данные прогресса
    if progress_data:
        # Используем переданные данные прогресса для продолжения
        current_index = progress_data.get('current_index', 0)
        found_results = progress_data.get('found_allowances', [])
        checked_wallets = set(progress_data.get('checked_wallets', []))
        zero_balance_count = progress_data.get('zero_balance_count', 0)
        api_errors = progress_data.get('api_errors', 0)
        
        # Фильтруем уже проверенные кошельки
        remaining_wallets = [w for i, w in enumerate(wallets) if i >= current_index]
        
        logger.info(f"🔄 ПРОДОЛЖЕНИЕ: продолжаем с {current_index}-го кошелька из {total_wallets}")
        start_from_scratch = False
        
    else:
        # Проверяем есть ли сохраненный прогресс в файле
        saved_progress = load_progress(user_id) if user_id else None
        
        if saved_progress and saved_progress.get('session_id') == session_id:
            # Продолжаем с места остановки из файла
            current_index = saved_progress['current_index']
            found_results = saved_progress['found_allowances']
            checked_wallets = set(saved_progress['checked_wallets'])
            zero_balance_count = saved_progress.get('zero_balance_count', 0)
            api_errors = saved_progress.get('api_errors', 0)
            
            # Фильтруем уже проверенные кошельки
            remaining_wallets = [w for i, w in enumerate(wallets) if i >= current_index]
            
            logger.info(f"🔄 Восстановлен прогресс: {current_index}/{total_wallets} кошельков")
            start_from_scratch = False
        else:
            # Начинаем новую проверку
            current_index = 0
            found_results = []
            checked_wallets = set()
            remaining_wallets = wallets
            zero_balance_count = 0
            api_errors = 0
            start_from_scratch = True
            
            # Сохраняем начальный прогресс
            if user_id:
                initial_progress = {
                    'session_id': session_id,
                    'user_id': user_id,
                    'total_wallets': total_wallets,
                    'current_index': 0,
                    'found_allowances': [],
                    'checked_wallets': [],
                    'failed_wallets': [],
                    'zero_balance_count': 0,
                    'api_errors': 0,
                    'start_time': datetime.now().isoformat(),
                    'spender': spender_address,
                    'token_contract': token_contract,
                    'network': network['name']
                }
                save_progress(user_id, initial_progress)
    
    processed_count = current_index
    
    logger.info(f"🛡️ [RELIABLE_SCAN] Начало проверки {len(remaining_wallets)} кошельков в {network['name']}")
    
    try:
        for i, wallet in enumerate(remaining_wallets):
            global_index = current_index + i
            
            # 🔧 ОПТИМИЗАЦИЯ: Периодически даем возможность обработать другие события
            if global_index % 20 == 0:
                await asyncio.sleep(0.05)
            
            try:
                # Обновляем прогресс каждые CACHE_SAVE_INTERVAL кошельков
                if global_index % CACHE_SAVE_INTERVAL == 0 and user_id:
                    save_progress(user_id, {
                        'session_id': session_id,
                        'user_id': user_id,
                        'total_wallets': total_wallets,
                        'current_index': global_index,
                        'found_allowances': found_results,
                        'checked_wallets': list(checked_wallets),
                        'failed_wallets': [],
                        'zero_balance_count': zero_balance_count,
                        'api_errors': api_errors,
                        'spender': spender_address,
                        'token_contract': token_contract,
                        'network': network['name']
                    })
                
                # Обновляем UI прогресс
                if progress_callback and i % 10 == 0:
                    progress = int((global_index / total_wallets) * 100)
                    status = (f"🔍 Проверка {global_index}/{total_wallets}\n"
                             f"✅ Найдено: {len(found_results)}\n"
                             f"💸 Нулевых: {zero_balance_count}")
                    await progress_callback(status, progress)
                
                # 🔧 ДЕБАГ: Логируем начало обработки кошелька
                logger.info(f"🔍 [WALLET_START] Обработка кошелька {i+1}/{len(remaining_wallets)}: {wallet[:8]}...")
                
                # 1. Проверяем баланс с retry
                is_eth = (token_contract == ETH_CONTRACT)
                logger.info(f"🔍 [BALANCE_CHECK] Проверяем баланс для {wallet[:8]}...")
                balance = await get_token_balance_with_retry(wallet, token_contract, network['chain_id'], is_eth)
                logger.info(f"🔍 [BALANCE_RESULT] {wallet[:8]}...: баланс = {balance:.6f}")
                
                if balance <= 0:
                    logger.info(f"🔍 [ZERO_BALANCE] {wallet[:8]}...: нулевой баланс, пропускаем")
                    zero_balance_count += 1
                    processed_count += 1
                    checked_wallets.add(wallet)
                    continue
                
                # 🔧 ДЕБАГ: Логируем начало проверки allowance
                logger.info(f"🔍 [ALLOWANCE_CHECK] Начинаем проверку allowance для {wallet[:8]}... в {network['name']} (chain_id: {network['chain_id']})")
                
                # 2. Проверяем allowance с retry - ДЛЯ ВСЕХ СЕТЕЙ!
                if is_eth:
                    allowance = balance  # Для ETH allowance = баланс
                    logger.info(f"🔍 [ETH_ALLOWANCE] {wallet[:8]}...: ETH, allowance = баланс = {balance:.6f}")
                else:
                    # 🔧 ДЛЯ ВСЕХ СЕТЕЙ ИСПОЛЬЗУЕМ ПРЯМОЙ ВЫЗОВ КОНТРАКТА ЧЕРЕЗ RPC
                    if network['chain_id'] == BASE_CHAIN_ID:
                        # Для Base используем Base RPC
                        allowance = await get_allowance_via_rpc(wallet, token_contract, spender_address)
                    elif network['chain_id'] == ARBITRUM_CHAIN_ID:
                        # Для Arbitrum используем Arbitrum RPC
                        correct_token_contract = USDC_CONTRACT_ARBITRUM  # 0xaf88d065e77c8cC2239327C5EDb3A432268e5831
                        allowance = await get_allowance_via_arbitrum_rpc(wallet, correct_token_contract, spender_address)
                    elif network['chain_id'] == OPTIMISM_CHAIN_ID:
                        # Для Optimism используем стандартный метод (пока)
                        allowance = await get_token_allowance(wallet, token_contract, spender_address, network['chain_id'])
                    else:
                        # Fallback - стандартный метод
                        allowance = await get_token_allowance(wallet, token_contract, spender_address, network['chain_id'])
                    logger.info(f"🔍 [USDC_ALLOWANCE_RESULT] {wallet[:8]}...: allowance = {allowance:.6f}")
                
                if allowance > 0:
                    result = {
                        'address': wallet,
                        'allowance': allowance,
                        'balance': balance,
                        'available': min(allowance, balance),
                        'network': network['name']
                    }
                    found_results.append(result)
                    logger.info(f"🎯 [SUCCESS] {wallet[:8]}...: {allowance:.6f} USDC (баланс: {balance:.6f})")
                else:
                    logger.info(f"🔍 [NO_ALLOWANCE] {wallet[:8]}...: allowance не найден или равен 0")
                
                processed_count += 1
                checked_wallets.add(wallet)
                
                # 🔧 ОПТИМИЗАЦИЯ: Уменьшенная пауза между кошельками
                pause_time = 0.3
                if i % 5 == 0:
                    await asyncio.sleep(pause_time)
                    
            except Exception as e:
                logger.error(f"❌ Критическая ошибка для {wallet[:8]}...: {e}")
                api_errors += 1
                processed_count += 1
                checked_wallets.add(wallet)
                continue
        
        # Финальная статистика
        logger.info(f"📊 [RELIABLE_SCAN] ИТОГО: "
                   f"Кошельков={total_wallets}, "
                   f"Обработано={processed_count}, "
                   f"Найдено={len(found_results)}, "
                   f"Нулевой баланс={zero_balance_count}, "
                   f"Ошибки={api_errors}")
        
        # Удаляем прогресс после успешного завершения
        if user_id:
            delete_progress(user_id)
        
        # Сортируем результаты
        found_results.sort(key=lambda x: x['allowance'], reverse=True)
        for i, result in enumerate(found_results):
            result['rank'] = i + 1
        
        return found_results
        
    except asyncio.CancelledError:
        logger.info("🔴 Проверка прервана пользователем")
        # Сохраняем прогресс при прерывании
        if user_id:
            save_progress(user_id, {
                'session_id': session_id,
                'user_id': user_id,
                'total_wallets': total_wallets,
                'current_index': processed_count,
                'found_allowances': found_results,
                'checked_wallets': list(checked_wallets),
                'failed_wallets': [],
                'zero_balance_count': zero_balance_count,
                'api_errors': api_errors,
                'spender': spender_address,
                'token_contract': token_contract,
                'network': network['name']
            })
        raise
        
    except Exception as e:
        logger.error(f"❌ Критическая ошибка в scan_allowance_reliable: {e}")
        # Сохраняем прогресс при ошибке
        if user_id:
            save_progress(user_id, {
                'session_id': session_id,
                'user_id': user_id,
                'total_wallets': total_wallets,
                'current_index': processed_count,
                'found_allowances': found_results,
                'checked_wallets': list(checked_wallets),
                'failed_wallets': [],
                'zero_balance_count': zero_balance_count,
                'api_errors': api_errors,
                'spender': spender_address,
                'token_contract': token_contract,
                'network': network['name']
            })
        raise
        
    except Exception as e:
        logger.error(f"❌ Критическая ошибка в scan_allowance_reliable: {e}")
        # Сохраняем прогресс при ошибке
        if user_id:
            save_progress(user_id, {
                'session_id': session_id,
                'user_id': user_id,
                'total_wallets': total_wallets,
                'current_index': processed_count,
                'found_allowances': found_results,
                'checked_wallets': list(checked_wallets),
                'failed_wallets': [],
                'zero_balance_count': zero_balance_count,
                'api_errors': api_errors,
                'spender': spender_address,
                'token_contract': token_contract,
                'network': network['name']
            })
        raise                        

@dp.message(Command("save_results"))
async def save_results_command(message: types.Message):
    """Срочное сохранение результатов"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    # Ищем последние результаты в глобальных переменных
    results_found = False
    
    # Проверяем current_allowance_data
    user_data = current_allowance_data.get(message.from_user.id, {})
    if user_data.get('last_results'):
        results = user_data['last_results']
        await export_and_send_results(message, results, "recovered", [])
        results_found = True
    
    # Проверяем contract_analysis_results
    user_sessions_list = [
        sid for sid in contract_analysis_results.keys() 
        if sid.startswith(str(message.from_user.id))
    ]
    
    if user_sessions_list and not results_found:
        last_session = user_sessions_list[-1]
        session_data = contract_analysis_results[last_session]
        wallets = session_data.get('wallets', [])
        
        if wallets:
            # Создаем фиктивные результаты для экспорта кошельков
            fake_results = [{'address': w, 'allowance': 0, 'balance': 0, 'available': 0, 'network': 'Base', 'rank': i+1} 
                          for i, w in enumerate(wallets)]
            
            txt_file = await export_wallets_to_txt(wallets, "recovered_wallets")
            if txt_file:
                await message.answer_document(
                    types.BufferedInputFile(
                        txt_file.getvalue(),
                        filename=f"recovered_wallets_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
                    ),
                    caption=f"✅ Восстановлено {len(wallets)} кошельков из последней сессии"
                )
                results_found = True
    
    if not results_found:
        await message.answer("❌ Не удалось найти результаты в памяти")



# ========== АВАРИЙНЫЕ КОМАНДЫ ДЛЯ СОХРАНЕНИЯ РЕЗУЛЬТАТОВ ==========

@dp.message(Command("emergency_export"))
async def emergency_export(message: types.Message):
    """Аварийный экспорт - создает файл с тестовыми данными на основе статистики"""
    if message.from_user.id not in ALLOWED_USERS:
        await message.answer("❌ Доступ запрещен")
        return
    
    # Создаем файл с примером на основе вашей статистики
    sample_results = []
    for i in range(124):
        sample_results.append({
            'address': f'0x{"{:040x}".format(i)}',  # Генерируем фиктивные адреса
            'allowance': 1000.0 + i * 10,
            'balance': 500.0 + i * 5,
            'available': 500.0 + i * 5,
            'network': 'Base',
            'rank': i + 1
        })
    
    excel_file = await export_to_excel(sample_results, "emergency_export")
    
    if excel_file:
        await message.answer_document(
            types.BufferedInputFile(
                excel_file.getvalue(),
                filename="EMERGENCY_allowance_results_template.xlsx"
            ),
            caption="⚠️ ШАБЛОН ДЛЯ РУЧНОГО ЗАПОЛНЕНИЯ\n\n"
                   "📊 Ваша статистика:\n"
                   "• ✅ Найдено: 124 кошелька\n"
                   "• 💸 Нулевых: 513\n"
                   "• 🔍 Проверено: 1224\n\n"
                   "💡 Сохраните этот файл и заполните реальными данными"
        )
    else:
        await message.answer("❌ Не удалось создать аварийный файл")

#ФУНКЦИЯ ЗАПУСКА ПРОВЕРКИ

async def start_allowance_check(user_id, network_choice, message=None):
    """Запускает проверку allowance с текущими параметрами"""
    try:
        user_data = current_allowance_data.get(user_id, {})
        wallets = user_data.get('wallets', [])
        token_type = user_data.get('token_type', 'usdc')
        spender_address = user_data.get('spender')
        progress_data = user_data.get('progress_data')  # Получаем данные прогресса
        
        if not wallets or not spender_address:
            if message:
                await message.answer("❌ Ошибка данных. Начните заново.")
            return
        
        token_contract = USDC_CONTRACT_BASE if token_type == 'usdc' else ETH_CONTRACT
        
        if message:
            progress_msg = await message.answer("🛡️ Запуск проверки allowance...")
        else:
            progress_msg = None
        
        async def progress_callback(text, progress):
            if progress_msg:
                try:
                    progress_bar = "█" * (progress // 10) + "░" * (10 - progress // 10)
                    await progress_msg.edit_text(f"{text}\n📊 Прогресс: {progress_bar} {progress}%")
                except Exception:
                    pass
        
        # Запускаем проверку
        if network_choice == 'all':
            all_results = []
            for network in NETWORK_PRIORITY:
                await progress_callback(f"🔍 Проверка в {network['name']}...", 0)
                network_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, 
                    user_id=user_id, progress_data=progress_data  # ← ПЕРЕДАЕМ ПРОГРЕСС
                )
                all_results.extend(network_results)
        else:
            network = next((n for n in NETWORK_PRIORITY if n['name'].lower() == network_choice), None)
            if network:
                # Передаем прогресс для продолжения
                all_results = await scan_allowance_reliable(
                    wallets, token_contract, spender_address, network, progress_callback, 
                    user_id=user_id, progress_data=progress_data  # ← ПЕРЕДАЕМ ПРОГРЕСС
                )   
            else:
                if message:
                    await message.answer("❌ Сеть не найдена")
                return
        
        # 📊 ФИНАЛЬНЫЙ ОТЧЕТ
        if all_results:
            result_text = (
                f"😈 **ПРОВЕРКА ЗАВЕРШЕНА**\n\n"
                f"📊 Статистика:\n"
                f"• 👛 Проверено кошельков: {len(wallets):,}\n"
                f"• ✅ Найдено allowance: {len(all_results):,}\n"
                f"• 💸 Нулевой баланс: {len(wallets) - len(all_results):,}\n"
            )
            
            # Экспорт результатов
            excel_file = await export_to_excel(all_results, f"allowance_{network_choice}")
            csv_file = await export_to_csv(all_results, f"allowance_{network_choice}")
            
            if excel_file:
                await message.answer_document(
                    types.BufferedInputFile(
                        excel_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            elif csv_file:
                await message.answer_document(
                    types.BufferedInputFile(
                        csv_file.getvalue(),
                        filename=f"allowance_results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
                    ),
                    caption=result_text,
                    parse_mode="Markdown"
                )
            else:
                await message.answer(result_text, parse_mode="Markdown")
                
        else:
            await message.answer(
                f"❌ **ALLOWANCE НЕ НАЙДЕН**\n\n"
                f"Проверено: {len(wallets):,} кошельков\n"
                f"Spender: `{spender_address}`\n"
                f"Токен: {token_type.upper()}\n"
                f"Сеть: {network_choice.title()}"
            )
        
    except Exception as e:
        logger.error(f"❌ Allowance check error: {e}")
        if message:
            await message.answer(f"❌ Ошибка проверки: {str(e)}")        
        
# ========== ЗАПУСК БОТА ==========

async def main():
    logger.info("🚀 Запуск бота-анализатора контрактов...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    print("🚀 Бот запускается на Render...")
    asyncio.run(main())